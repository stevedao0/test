from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Ensure project root is on sys.path when running as a script
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.config import STORAGE_EXCEL_DIR  # noqa: E402
from app.db import DB_PATH, engine, session_scope  # noqa: E402
from app.db_models import Base, ContractRecordRow, WorkRow  # noqa: E402
from app.services.excel_store import read_contracts  # noqa: E402


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s)
        except Exception:
            return None
    return None


def _to_int(v):
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip().replace(".", "").replace(",", "")
        if not s:
            return None
        try:
            return int(s)
        except Exception:
            return None
    return None


def migrate_contracts() -> int:
    count = 0
    for p in sorted(STORAGE_EXCEL_DIR.glob("contracts_*.xlsx")):
        year = None
        try:
            name = p.stem
            if name.startswith("contracts_"):
                year = int(name.split("_")[-1])
        except Exception:
            year = None

        rows = read_contracts(excel_path=p)
        for r in rows:
            # Normalize keys to our db schema
            row = ContractRecordRow(
                contract_no=(r.get("contract_no") or ""),
                contract_year=int(r.get("contract_year") or year or 0),
                annex_no=(r.get("annex_no") or None) or None,
                ngay_lap_hop_dong=_to_date(r.get("ngay_lap_hop_dong")),
                linh_vuc=r.get("linh_vuc"),
                region_code=r.get("region_code"),
                field_code=r.get("field_code"),
                don_vi_ten=r.get("don_vi_ten"),
                don_vi_dia_chi=r.get("don_vi_dia_chi"),
                don_vi_dien_thoai=r.get("don_vi_dien_thoai"),
                don_vi_nguoi_dai_dien=r.get("don_vi_nguoi_dai_dien"),
                don_vi_chuc_vu=r.get("don_vi_chuc_vu"),
                don_vi_mst=r.get("don_vi_mst"),
                don_vi_email=r.get("don_vi_email"),
                so_cccd=r.get("so_CCCD") or r.get("so_cccd"),
                ngay_cap_cccd=r.get("ngay_cap_CCCD") or r.get("ngay_cap_cccd"),
                kenh_ten=r.get("kenh_ten"),
                kenh_id=r.get("kenh_id"),
                nguoi_thuc_hien_email=r.get("nguoi_thuc_hien_email"),
                so_tien_nhuan_but_value=_to_int(r.get("so_tien_nhuan_but_value")),
                so_tien_nhuan_but_text=r.get("so_tien_nhuan_but_text"),
                so_tien_chua_gtgt_value=_to_int(r.get("so_tien_chua_GTGT_value")),
                so_tien_chua_gtgt_text=r.get("so_tien_chua_GTGT_text"),
                thue_percent=(float(r.get("thue_percent")) if r.get("thue_percent") not in (None, "") else None),
                thue_gtgt_value=_to_int(r.get("thue_GTGT_value")),
                thue_gtgt_text=r.get("thue_GTGT_text"),
                so_tien_value=_to_int(r.get("so_tien_value")),
                so_tien_text=r.get("so_tien_text"),
                so_tien_bang_chu=r.get("so_tien_bang_chu"),
                docx_path=r.get("docx_path"),
                catalogue_path=r.get("catalogue_path"),
            )

            # Skip obviously bad rows
            if not row.contract_no or not row.contract_year:
                continue

            try:
                with session_scope() as db:
                    db.add(row)
                count += 1
            except Exception:
                # Likely duplicate uq; ignore and continue
                continue

    return count


def migrate_works() -> int:
    count = 0
    for p in sorted(STORAGE_EXCEL_DIR.glob("works_contract_*.xlsx")):
        wb = load_workbook(str(p), data_only=True)
        ws = wb["Works"] if "Works" in wb.sheetnames else wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            wb.close()
            continue

        headers = [h if isinstance(h, str) else "" for h in list(values[0])]
        header_map = {h: i for i, h in enumerate(headers) if h}

        for row_vals in values[1:]:
            if not any(row_vals):
                continue

            def g(key: str):
                idx = header_map.get(key)
                return row_vals[idx] if idx is not None and idx < len(row_vals) else None

            rec = WorkRow(
                year=int(g("year") or 0),
                contract_no=str(g("contract_no") or ""),
                annex_no=(str(g("annex_no") or "").strip() or None),
                ngay_ky_hop_dong=str(g("ngay_ky_hop_dong") or ""),
                ngay_ky_phu_luc=str(g("ngay_ky_phu_luc") or ""),
                nguoi_thuc_hien=str(g("nguoi_thuc_hien") or ""),
                ten_kenh=str(g("ten_kenh") or ""),
                id_channel=str(g("id_channel") or ""),
                link_kenh=str(g("link_kenh") or ""),
                stt=_to_int(g("stt")),
                id_link=str(g("id_link") or ""),
                youtube_url=str(g("youtube_url") or ""),
                id_work=str(g("id_work") or ""),
                musical_work=str(g("musical_work") or ""),
                author=str(g("author") or ""),
                composer=str(g("composer") or ""),
                lyricist=str(g("lyricist") or ""),
                time_range=str(g("time_range") or ""),
                duration=str(g("duration") or ""),
                effective_date=str(g("effective_date") or ""),
                expiration_date=str(g("expiration_date") or ""),
                usage_type=str(g("usage_type") or ""),
                royalty_rate=str(g("royalty_rate") or ""),
                note=str(g("note") or ""),
                imported_at=str(g("imported_at") or ""),
            )

            if not rec.year or not rec.contract_no:
                continue

            try:
                with session_scope() as db:
                    db.add(rec)
                count += 1
            except Exception:
                continue

        wb.close()

    return count


def main() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    Base.metadata.create_all(bind=engine)

    c = migrate_contracts()
    w = migrate_works()

    print(f"DB: {DB_PATH}")
    print(f"Imported contracts rows: {c}")
    print(f"Imported works rows: {w}")


if __name__ == "__main__":
    main()
