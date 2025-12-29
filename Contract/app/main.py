from __future__ import annotations

import re
import traceback
import os
import hashlib
import hmac
import secrets
from io import BytesIO
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, Response
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from starlette import status
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.config import (
    ANNEX_TEMPLATE_PATH,
    ANNEX_CATALOGUE_TEMPLATE_PATH,
    CATALOGUE_TEMPLATE_PATH,
    DOCX_TEMPLATE_PATH,
    STORAGE_DIR,
    STORAGE_DOCX_DIR,
    STORAGE_EXCEL_DIR,
    UI_STATIC_DIR,
    UI_TEMPLATES_DIR,
    WEB_TEMPLATES_DIR,
)
from app.documents.naming import build_docx_filename
from app.models import ContractCreate, ContractRecord
# from app.services.annex_store import append_annex_row  # No longer needed - annexes saved to contracts Excel
from app.services.docx_renderer import date_parts, render_contract_docx
from app.services.excel_store import (
    export_catalogue_excel,
    read_contracts,
)
from app.services.safety import audit_log, safe_move_to_backup, safe_replace_bytes

from app.services.excel_store import HEADERS, WORKS_HEADERS

from app.db import DB_PATH, engine, session_scope
from app.db_models import Base, ContractRecordRow, UserRow, WorkRow


app = FastAPI()


_basic_auth = HTTPBasic()


def _hash_password(password: str, *, salt_hex: str, iterations: int = 200_000) -> str:
    dk = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        bytes.fromhex(salt_hex),
        iterations,
    )
    return dk.hex()


def _verify_password(password: str, *, salt_hex: str, expected_hash_hex: str) -> bool:
    calc = _hash_password(password, salt_hex=salt_hex)
    return hmac.compare_digest(calc, expected_hash_hex)


def _ensure_default_users() -> None:
    admin_pwd = os.environ.get("CONTRACT_ADMIN_PASSWORD")
    mod_pwd = os.environ.get("CONTRACT_MOD_PASSWORD")

    if not admin_pwd:
        admin_pwd = "admin123"
    if not mod_pwd:
        mod_pwd = "mod123"

    with session_scope() as db:
        has_any = db.query(UserRow).first() is not None
        if has_any:
            return

        admin_salt = secrets.token_bytes(16).hex()
        mod_salt = secrets.token_bytes(16).hex()

        db.add(
            UserRow(
                username="admin",
                role="admin",
                password_salt=admin_salt,
                password_hash=_hash_password(admin_pwd, salt_hex=admin_salt),
            )
        )
        db.add(
            UserRow(
                username="mod",
                role="mod",
                password_salt=mod_salt,
                password_hash=_hash_password(mod_pwd, salt_hex=mod_salt),
            )
        )


def _get_current_user(credentials: HTTPBasicCredentials = Depends(_basic_auth)) -> UserRow:
    with session_scope() as db:
        user = db.query(UserRow).filter(UserRow.username == credentials.username).first()

    if user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )

    if not _verify_password(credentials.password, salt_hex=user.password_salt, expected_hash_hex=user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )

    return user


def require_role(*allowed_roles: str):
    def _dep(user: UserRow = Depends(_get_current_user)) -> UserRow:
        if user.role not in allowed_roles:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
        return user

    return _dep


@app.on_event("startup")
def _startup_db() -> None:
    # Ensure SQLite schema exists
    Base.metadata.create_all(bind=engine)
    _ensure_default_users()


_BACKUPS_DIR = STORAGE_DIR / "backups"
_LOGS_DIR = STORAGE_DIR / "logs"


def _db_available() -> bool:
    try:
        return DB_PATH.exists()
    except Exception:
        return False


def _rows_from_db(*, year: int) -> list[dict]:
    if not _db_available():
        return []

    with session_scope() as db:
        q = db.query(ContractRecordRow).filter(ContractRecordRow.contract_year == year)
        out: list[dict] = []
        for r in q.all():
            out.append(
                {
                    "contract_no": r.contract_no,
                    "contract_year": r.contract_year,
                    "annex_no": r.annex_no,
                    "ngay_lap_hop_dong": r.ngay_lap_hop_dong,
                    "linh_vuc": r.linh_vuc,
                    "region_code": r.region_code,
                    "field_code": r.field_code,
                    "don_vi_ten": r.don_vi_ten,
                    "don_vi_dia_chi": r.don_vi_dia_chi,
                    "don_vi_dien_thoai": r.don_vi_dien_thoai,
                    "don_vi_nguoi_dai_dien": r.don_vi_nguoi_dai_dien,
                    "don_vi_chuc_vu": r.don_vi_chuc_vu,
                    "don_vi_mst": r.don_vi_mst,
                    "don_vi_email": r.don_vi_email,
                    "so_CCCD": r.so_cccd,
                    "ngay_cap_CCCD": r.ngay_cap_cccd,
                    "kenh_ten": r.kenh_ten,
                    "kenh_id": r.kenh_id,
                    "nguoi_thuc_hien_email": r.nguoi_thuc_hien_email,
                    "so_tien_nhuan_but_value": r.so_tien_nhuan_but_value,
                    "so_tien_nhuan_but_text": r.so_tien_nhuan_but_text,
                    "so_tien_chua_GTGT_value": r.so_tien_chua_gtgt_value,
                    "so_tien_chua_GTGT_text": r.so_tien_chua_gtgt_text,
                    "thue_percent": r.thue_percent,
                    "thue_GTGT_value": r.thue_gtgt_value,
                    "thue_GTGT_text": r.thue_gtgt_text,
                    "so_tien_value": r.so_tien_value,
                    "so_tien_text": r.so_tien_text,
                    "so_tien_bang_chu": r.so_tien_bang_chu,
                    "docx_path": r.docx_path,
                    "catalogue_path": r.catalogue_path,
                }
            )
        return out


def _db_get_contract_row(*, year: int, contract_no: str, annex_no: str | None) -> ContractRecordRow | None:
    with session_scope() as db:
        return (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
            .filter(ContractRecordRow.annex_no.is_(annex_no) if annex_no is None else (ContractRecordRow.annex_no == annex_no))
            .first()
        )


def _db_upsert_contract_record(*, record: dict) -> None:
    year = int(record.get("contract_year") or 0)
    contract_no = str(record.get("contract_no") or "")
    annex_no = record.get("annex_no")
    annex_no = (str(annex_no).strip() if annex_no is not None else None) or None

    with session_scope() as db:
        q = (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
        )
        if annex_no is None:
            q = q.filter(ContractRecordRow.annex_no.is_(None))
        else:
            q = q.filter(ContractRecordRow.annex_no == annex_no)

        row = q.first()
        if row is None:
            row = ContractRecordRow(contract_year=year, contract_no=contract_no, annex_no=annex_no)
            db.add(row)

        # Map dict -> columns
        for k, v in record.items():
            if k in ("so_CCCD", "ngay_cap_CCCD"):
                continue
            if hasattr(row, k):
                setattr(row, k, v)

        # Special casing for legacy key casing
        if "so_CCCD" in record:
            row.so_cccd = record.get("so_CCCD")
        if "ngay_cap_CCCD" in record:
            row.ngay_cap_cccd = record.get("ngay_cap_CCCD")


def _db_update_contract_fields(*, year: int, contract_no: str, annex_no: str | None, updated: dict) -> bool:
    annex_no = (annex_no.strip() if isinstance(annex_no, str) else annex_no) or None

    with session_scope() as db:
        q = (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
        )
        if annex_no is None:
            q = q.filter(ContractRecordRow.annex_no.is_(None))
        else:
            q = q.filter(ContractRecordRow.annex_no == annex_no)

        row = q.first()
        if row is None:
            return False

        for k, v in updated.items():
            if k in ("so_CCCD", "ngay_cap_CCCD"):
                continue
            if hasattr(row, k):
                setattr(row, k, v)

        if "so_CCCD" in updated:
            row.so_cccd = updated.get("so_CCCD")
        if "ngay_cap_CCCD" in updated:
            row.ngay_cap_cccd = updated.get("ngay_cap_CCCD")

        return True


def _db_delete_contract_record(*, year: int, contract_no: str, annex_no: str | None) -> bool:
    annex_no = (annex_no.strip() if isinstance(annex_no, str) else annex_no) or None
    with session_scope() as db:
        q = (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
        )
        if annex_no is None:
            q = q.filter(ContractRecordRow.annex_no.is_(None))
        else:
            q = q.filter(ContractRecordRow.annex_no == annex_no)

        row = q.first()
        if row is None:
            return False
        db.delete(row)
        return True


_HEADER_FONT = Font(bold=True)


def _xlsx_bytes_from_rows(*, sheet_name: str, headers: list[str], rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT

    for r_idx, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c, value=row.get(h))

    bio = BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue()


def _export_contracts_excel_bytes(*, year: int) -> bytes:
    # Build rows matching legacy Excel headers
    with session_scope() as db:
        q = db.query(ContractRecordRow).filter(ContractRecordRow.contract_year == year)
        db_rows = q.all()

    rows: list[dict] = []
    for r in db_rows:
        rows.append(
            {
                "contract_no": r.contract_no,
                "contract_year": r.contract_year,
                "annex_no": r.annex_no,
                "ngay_lap_hop_dong": r.ngay_lap_hop_dong,
                "linh_vuc": r.linh_vuc,
                "region_code": r.region_code,
                "field_code": r.field_code,
                "don_vi_ten": r.don_vi_ten,
                "don_vi_dia_chi": r.don_vi_dia_chi,
                "don_vi_dien_thoai": r.don_vi_dien_thoai,
                "don_vi_nguoi_dai_dien": r.don_vi_nguoi_dai_dien,
                "don_vi_chuc_vu": r.don_vi_chuc_vu,
                "don_vi_mst": r.don_vi_mst,
                "don_vi_email": r.don_vi_email,
                "so_CCCD": r.so_cccd,
                "ngay_cap_CCCD": r.ngay_cap_cccd,
                "kenh_ten": r.kenh_ten,
                "kenh_id": r.kenh_id,
                "nguoi_thuc_hien_email": r.nguoi_thuc_hien_email,
                "so_tien_nhuan_but_value": r.so_tien_nhuan_but_value,
                "so_tien_nhuan_but_text": r.so_tien_nhuan_but_text,
                "so_tien_chua_GTGT_value": r.so_tien_chua_gtgt_value,
                "so_tien_chua_GTGT_text": r.so_tien_chua_gtgt_text,
                "thue_percent": r.thue_percent,
                "thue_GTGT_value": r.thue_gtgt_value,
                "thue_GTGT_text": r.thue_gtgt_text,
                "so_tien_value": r.so_tien_value,
                "so_tien_text": r.so_tien_text,
                "so_tien_bang_chu": r.so_tien_bang_chu,
                "docx_path": r.docx_path,
                "catalogue_path": r.catalogue_path,
            }
        )

    return _xlsx_bytes_from_rows(sheet_name="Contracts", headers=list(HEADERS), rows=rows)


def _export_works_excel_bytes(*, year: int) -> bytes:
    with session_scope() as db:
        q = db.query(WorkRow).filter(WorkRow.year == year)
        db_rows = q.all()

    rows: list[dict] = []
    for r in db_rows:
        rows.append(
            {
                "year": r.year,
                "contract_no": r.contract_no,
                "annex_no": r.annex_no,
                "ngay_ky_hop_dong": r.ngay_ky_hop_dong,
                "ngay_ky_phu_luc": r.ngay_ky_phu_luc,
                "nguoi_thuc_hien": r.nguoi_thuc_hien,
                "ten_kenh": r.ten_kenh,
                "id_channel": r.id_channel,
                "link_kenh": r.link_kenh,
                "stt": r.stt,
                "id_link": r.id_link,
                "youtube_url": r.youtube_url,
                "id_work": r.id_work,
                "musical_work": r.musical_work,
                "author": r.author,
                "composer": r.composer,
                "lyricist": r.lyricist,
                "time_range": r.time_range,
                "duration": r.duration,
                "effective_date": r.effective_date,
                "expiration_date": r.expiration_date,
                "usage_type": r.usage_type,
                "royalty_rate": r.royalty_rate,
                "note": r.note,
                "imported_at": r.imported_at,
            }
        )

    return _xlsx_bytes_from_rows(sheet_name="Works", headers=list(WORKS_HEADERS), rows=rows)


@app.get("/debug/contracts")
def debug_contracts(year: int | None = None):
    y = _pick_year(year)
    rows = _rows_from_db(year=y) if _db_available() else []
    contracts = [r for r in rows if not r.get("annex_no")]
    annexes = [r for r in rows if r.get("annex_no")]
    sample = contracts[0] if contracts else (rows[0] if rows else None)
    return JSONResponse(
        {
            "year": y,
            "db_path": str(DB_PATH),
            "db_exists": _db_available(),
            "rows": len(rows),
            "contracts": len(contracts),
            "annexes": len(annexes),
            "sample": sample,
        }
    )


@app.get("/catalogue/upload", response_class=HTMLResponse)
def catalogue_upload_form(
    request: Request,
    year: int | None = None,
    contract_no: str | None = None,
    annex_no: str | None = None,
    error: str | None = None,
    message: str | None = None,
):
    y = _pick_year(year or (_year_from_contract_no(contract_no or "") if contract_no else None))
    return templates.TemplateResponse(
        "catalogue_upload.html",
        {
            "request": request,
            "title": "Upload danh mục Excel",
            "year": y,
            "contract_no": contract_no or "",
            "annex_no": annex_no or "",
            "error": error,
            "message": message,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/catalogue/upload")
async def catalogue_upload_submit(
    request: Request,
    year: int = Form(...),
    contract_no: str = Form(...),
    annex_no: str = Form(""),
    catalogue_file: UploadFile = File(...),
    user: UserRow = Depends(require_role("admin", "mod")),
):
    try:
        if not catalogue_file.filename or not catalogue_file.filename.lower().endswith(".xlsx"):
            raise ValueError("File danh mục phải là .xlsx")

        target_annex_no = annex_no.strip() or None
        if _db_get_contract_row(year=year, contract_no=contract_no, annex_no=target_annex_no) is None:
            raise ValueError("Không tìm thấy hợp đồng/phụ lục để cập nhật catalogue_path")

        data = await catalogue_file.read()
        if data is None or len(data) == 0:
            raise ValueError("File upload rỗng")
        if len(data) > 25 * 1024 * 1024:
            raise ValueError("File upload quá lớn (tối đa 25MB)")

        out_dir = STORAGE_EXCEL_DIR / str(year)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / Path(catalogue_file.filename).name
        safe_replace_bytes(out_path, data, backup_dir=_BACKUPS_DIR / "files")

        success = _db_update_contract_fields(
            year=year,
            contract_no=contract_no,
            annex_no=target_annex_no,
            updated={"catalogue_path": str(out_path)},
        )
        if not success:
            raise ValueError("Không tìm thấy hợp đồng/phụ lục để cập nhật catalogue_path")

        audit_log(
            log_dir=_LOGS_DIR,
            event={
                "action": "catalogue.upload",
                "ip": getattr(getattr(request, "client", None), "host", None),
                "year": year,
                "contract_no": contract_no,
                "annex_no": target_annex_no,
                "file": out_path.name,
                "bytes": len(data),
            },
        )

        return RedirectResponse(
            url=f"/catalogue/upload?year={year}&contract_no={contract_no}&annex_no={annex_no}&message=Đã upload danh mục và cập nhật dữ liệu",
            status_code=303,
        )
    except Exception as e:
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(
            url=f"/catalogue/upload?year={year}&contract_no={contract_no}&annex_no={annex_no}&error={msg}",
            status_code=303,
        )


def _pick_year(year: int | None) -> int:
    if year:
        return year

    try:
        years: list[int] = []
        for p in STORAGE_EXCEL_DIR.glob("contracts_*.xlsx"):
            m = re.match(r"contracts_(\d{4})\.xlsx$", p.name)
            if m:
                years.append(int(m.group(1)))
        if years:
            return max(years)
    except Exception:
        pass

    return date.today().year


@app.get("/")
def home() -> RedirectResponse:
    return RedirectResponse(url="/documents/new")


@app.get("/contracts/new")
def contract_form() -> RedirectResponse:
    return RedirectResponse(url="/documents/new?doc_type=contract")


@app.get("/annexes/new")
def annex_form() -> RedirectResponse:
    return RedirectResponse(url="/documents/new?doc_type=annex")


@app.get("/api/contracts")
def api_contracts_list(year: int | None = None, q: str | None = None):
    y = _pick_year(year)
    rows = _rows_from_db(year=y)
    contracts = [r for r in rows if not r.get("annex_no")]
    if q:
        ql = q.lower()
        contracts = [
            c
            for c in contracts
            if ql in (c.get("contract_no") or "").lower() or ql in (c.get("kenh_ten") or "").lower()
        ]
    result = []
    for c in contracts:
        result.append(
            {
                "contract_no": c.get("contract_no"),
                "kenh_ten": c.get("kenh_ten"),
                "don_vi_ten": c.get("don_vi_ten"),
                "kenh_id": c.get("kenh_id"),
            }
        )
    return JSONResponse({"contracts": result})


@app.get("/contracts", response_class=HTMLResponse)
def contracts_list(request: Request, response: Response, year: int | None = None, download: str | None = None, download2: str | None = None):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    y = _pick_year(year)
    rows = _rows_from_db(year=y)
    contracts = [r for r in rows if not r.get("annex_no")]

    catalogue_filter = (request.query_params.get("catalogue") or "all").strip().lower()
    if catalogue_filter in ("yes", "has", "1", "true"):
        contracts = [r for r in contracts if r.get("catalogue_path")]
    elif catalogue_filter in ("no", "none", "0", "false"):
        contracts = [r for r in contracts if not r.get("catalogue_path")]

    annexes = [r for r in rows if r.get("annex_no")]
    for r in contracts:
        contract_no = r.get("contract_no")
        r["annex_count"] = len([a for a in annexes if a.get("contract_no") == contract_no])

        p = Path(r.get("docx_path") or "")
        r["download_url"] = f"/download/{y}/{p.name}" if p.exists() else None

        cp = Path(r.get("catalogue_path") or "")
        r["catalogue_download_url"] = f"/download_excel/{y}/{cp.name}" if cp.exists() else None

    stats = {
        "total_contracts": len(contracts),
        "total_value": sum(int(r.get("so_tien_value") or 0) for r in contracts),
        "contracts_with_annexes": len({a.get("contract_no") for a in annexes if a.get("contract_no")}),
    }

    return templates.TemplateResponse(
        "contracts_list.html",
        {
            "request": request,
            "title": "Danh sách hợp đồng",
            "year": y,
            "rows": contracts,
            "stats": stats,
            "download": download,
            "download2": download2,
            "catalogue_filter": catalogue_filter,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.get("/annexes", response_class=HTMLResponse)
def annexes_list(request: Request, response: Response, year: int | None = None, download: str | None = None):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    y = _pick_year(year)
    rows = _rows_from_db(year=y)
    annexes = [r for r in rows if r.get("annex_no")]

    catalogue_filter = (request.query_params.get("catalogue") or "all").strip().lower()
    if catalogue_filter in ("yes", "has", "1", "true"):
        annexes = [r for r in annexes if r.get("catalogue_path")]
    elif catalogue_filter in ("no", "none", "0", "false"):
        annexes = [r for r in annexes if not r.get("catalogue_path")]

    contracts = [r for r in rows if not r.get("annex_no")]
    contracts_map = {r.get("contract_no"): r for r in contracts}

    for r in annexes:
        p = Path(r.get("docx_path") or "")
        r["download_url"] = f"/download/{y}/{p.name}" if p.exists() else None

        cp = Path(r.get("catalogue_path") or "")
        r["catalogue_download_url"] = f"/download_excel/{y}/{cp.name}" if cp.exists() else None

        parent = contracts_map.get(r.get("contract_no"))
        if parent:
            r["parent_contract"] = {
                "don_vi_ten": parent.get("don_vi_ten", ""),
                "kenh_ten": parent.get("kenh_ten", ""),
                "ngay_lap_hop_dong": parent.get("ngay_lap_hop_dong", ""),
            }
        else:
            r["parent_contract"] = None

    stats = {
        "total_annexes": len(annexes),
        "total_value": sum(int(r.get("so_tien_value") or 0) for r in annexes),
    }

    return templates.TemplateResponse(
        "annexes_list.html",
        {
            "request": request,
            "title": "Danh sách phụ lục",
            "year": y,
            "rows": annexes,
            "stats": stats,
            "download": download,
            "catalogue_filter": catalogue_filter,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/contracts")
def create_contract(
    request: Request,
    ngay_lap_hop_dong: str = Form(...),
    so_hop_dong_4: str = Form(...),
    linh_vuc: str = Form("Sao chép trực tuyến"),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form("Giám đốc"),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    so_CCCD: str = Form(""),
    ngay_cap_CCCD: str = Form(""),
    nguoi_thuc_hien_email: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form(""),
    user: UserRow = Depends(require_role("admin", "mod")),
):
    try:
        channel_id, channel_link = normalize_youtube_channel_input(kenh_id)
        linh_vuc_value = _clean_opt(linh_vuc) or "Sao chép trực tuyến"

        pre_vat_value: Optional[int] = None
        vat_percent_value: Optional[float] = None
        vat_value: Optional[int] = None
        total_value: Optional[int] = None

        if _clean_opt(so_tien_chua_GTGT):
            pre_vat_value = normalize_money_to_int(_clean_opt(so_tien_chua_GTGT))
            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_value = float(pct_raw.replace(",", "."))
            vat_value = int(round(pre_vat_value * vat_percent_value / 100.0))
            total_value = pre_vat_value + vat_value

        contract_date = date.fromisoformat(ngay_lap_hop_dong)
        year = contract_date.year
        contract_no = f"{so_hop_dong_4}/{year}/{REGION_CODE}/{FIELD_CODE}"

        if _db_get_contract_row(year=year, contract_no=contract_no, annex_no=None) is not None:
            raise ValueError("Số hợp đồng đã tồn tại")

        out_docx_dir = STORAGE_DOCX_DIR / str(year)
        out_docx_dir.mkdir(parents=True, exist_ok=True)
        filename = build_docx_filename(
            year=year,
            so_hop_dong_4=so_hop_dong_4,
            so_phu_luc=None,
            linh_vuc=linh_vuc_value,
            kenh_ten=_clean_opt(kenh_ten),
        )
        out_docx_path = out_docx_dir / filename
        if out_docx_path.exists():
            stem = out_docx_path.stem
            out_docx_path = out_docx_dir / f"{stem}_{date.today().strftime('%Y%m%d')}.docx"

        context = {
            "contract_no": contract_no,
            "so_hop_dong": contract_no,
            "linh_vuc": linh_vuc_value,
            **date_parts(contract_date),
            "don_vi_ten": _clean_opt(don_vi_ten),
            "don_vi_dia_chi": _clean_opt(don_vi_dia_chi),
            "don_vi_dien_thoai": normalize_multi_phones(don_vi_dien_thoai),
            "don_vi_nguoi_dai_dien": _clean_opt(don_vi_nguoi_dai_dien),
            "don_vi_chuc_vu": _clean_opt(don_vi_chuc_vu) or "Giám đốc",
            "don_vi_mst": _clean_opt(don_vi_mst),
            "don_vi_email": normalize_multi_emails(don_vi_email),
            "so_CCCD": _clean_opt(so_CCCD),
            "ngay_cap_CCCD": _clean_opt(ngay_cap_CCCD),
            "kenh_ten": _clean_opt(kenh_ten),
            "kenh_id": channel_id,
            "link_kenh": channel_link,
            "nguoi_thuc_hien_email": normalize_multi_emails(nguoi_thuc_hien_email),
            "so_tien_chua_GTGT": format_money_number(pre_vat_value) if pre_vat_value else "",
            "thue_GTGT": format_money_number(vat_value) if vat_value else "",
            "so_tien": format_money_number(total_value) if total_value else "",
            "so_tien_bang_chu": money_to_vietnamese_words(total_value) if total_value else "",
            "thue_percent": str(int(vat_percent_value)) if vat_percent_value else "10",
        }

        render_contract_docx(template_path=DOCX_TEMPLATE_PATH, output_path=out_docx_path, context=context)

        out_excel_dir = STORAGE_EXCEL_DIR / str(year)
        out_excel_dir.mkdir(parents=True, exist_ok=True)
        out_catalogue_path = out_excel_dir / out_docx_path.with_suffix(".xlsx").name
        export_catalogue_excel(
            template_path=CATALOGUE_TEMPLATE_PATH,
            output_path=out_catalogue_path,
            context=dict(context),
            sheet_name="Final",
        )

        _db_upsert_contract_record(
            record={
                "contract_no": contract_no,
                "contract_year": year,
                "annex_no": None,
                "ngay_lap_hop_dong": contract_date,
                "linh_vuc": linh_vuc_value,
                "region_code": REGION_CODE,
                "field_code": FIELD_CODE,
                "don_vi_ten": _clean_opt(don_vi_ten),
                "don_vi_dia_chi": _clean_opt(don_vi_dia_chi),
                "don_vi_dien_thoai": normalize_multi_phones(don_vi_dien_thoai),
                "don_vi_nguoi_dai_dien": _clean_opt(don_vi_nguoi_dai_dien),
                "don_vi_chuc_vu": _clean_opt(don_vi_chuc_vu) or "Giám đốc",
                "don_vi_mst": _clean_opt(don_vi_mst),
                "don_vi_email": normalize_multi_emails(don_vi_email),
                "so_CCCD": _clean_opt(so_CCCD),
                "ngay_cap_CCCD": _clean_opt(ngay_cap_CCCD),
                "kenh_ten": _clean_opt(kenh_ten),
                "kenh_id": channel_id,
                "nguoi_thuc_hien_email": normalize_multi_emails(nguoi_thuc_hien_email),
                "so_tien_chua_GTGT_value": pre_vat_value,
                "so_tien_chua_GTGT_text": format_money_number(pre_vat_value) if pre_vat_value else "",
                "thue_percent": vat_percent_value,
                "thue_GTGT_value": vat_value,
                "thue_GTGT_text": format_money_number(vat_value) if vat_value else "",
                "so_tien_value": total_value,
                "so_tien_text": format_money_number(total_value) if total_value else "",
                "so_tien_bang_chu": money_to_vietnamese_words(total_value) if total_value else "",
                "docx_path": str(out_docx_path),
                "catalogue_path": str(out_catalogue_path),
            }
        )

        audit_log(
            log_dir=_LOGS_DIR,
            event={
                "action": "contracts.create",
                "ip": getattr(getattr(request, "client", None), "host", None),
                "year": year,
                "contract_no": contract_no,
                "actor": user.username,
            },
        )

        return RedirectResponse(
            url=(
                f"/contracts?year={year}"
                f"&download=/download/{year}/{out_docx_path.name}"
                f"&download2=/download_excel/{year}/{out_catalogue_path.name}"
            ),
            status_code=303,
        )
    except Exception as e:
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(url=f"/documents/new?doc_type=contract&error={msg}", status_code=303)


@app.post("/contracts/{year}/delete")
def delete_contract(request: Request, year: int, contract_no: str, user: UserRow = Depends(require_role("admin", "mod"))):
    try:
        row = _db_get_contract_row(year=year, contract_no=contract_no, annex_no=None)
        if row and row.docx_path:
            p = Path(row.docx_path)
            if p.exists():
                safe_move_to_backup(p, backup_dir=_BACKUPS_DIR / "deleted")
        if row and row.catalogue_path:
            p = Path(row.catalogue_path)
            if p.exists():
                safe_move_to_backup(p, backup_dir=_BACKUPS_DIR / "deleted")

        ok = _db_delete_contract_record(year=year, contract_no=contract_no, annex_no=None)
        if ok:
            audit_log(
                log_dir=_LOGS_DIR,
                event={
                    "action": "contracts.delete",
                    "ip": getattr(getattr(request, "client", None), "host", None),
                    "year": year,
                    "contract_no": contract_no,
                    "actor": user.username,
                },
            )
            return JSONResponse({"success": True})
        return JSONResponse({"success": False, "error": "Không tìm thấy hợp đồng"}, status_code=404)
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/annexes/{year}/delete")
def delete_annex(request: Request, year: int, contract_no: str, annex_no: str, user: UserRow = Depends(require_role("admin", "mod"))):
    try:
        row = _db_get_contract_row(year=year, contract_no=contract_no, annex_no=annex_no)
        if row and row.docx_path:
            p = Path(row.docx_path)
            if p.exists():
                safe_move_to_backup(p, backup_dir=_BACKUPS_DIR / "deleted")
        if row and row.catalogue_path:
            p = Path(row.catalogue_path)
            if p.exists():
                safe_move_to_backup(p, backup_dir=_BACKUPS_DIR / "deleted")

        ok = _db_delete_contract_record(year=year, contract_no=contract_no, annex_no=annex_no)
        if ok:
            audit_log(
                log_dir=_LOGS_DIR,
                event={
                    "action": "annexes.delete",
                    "ip": getattr(getattr(request, "client", None), "host", None),
                    "year": year,
                    "contract_no": contract_no,
                    "annex_no": annex_no,
                    "actor": user.username,
                },
            )
            return JSONResponse({"success": True})
        return JSONResponse({"success": False, "error": "Không tìm thấy phụ lục"}, status_code=404)
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.get("/admin/ops", response_class=HTMLResponse)
def admin_ops(request: Request, user: UserRow = Depends(require_role("admin"))):
    logs_dir = STORAGE_DIR / "logs"
    backups_dir = STORAGE_DIR / "backups"

    logs: list[dict] = []
    backups: list[dict] = []

    if logs_dir.exists():
        for p in sorted(logs_dir.glob("**/*")):
            if p.is_file():
                logs.append({"name": str(p.relative_to(logs_dir)).replace("\\", "/"), "size": p.stat().st_size})

    if backups_dir.exists():
        for p in sorted(backups_dir.glob("**/*")):
            if p.is_file():
                backups.append({"name": str(p.relative_to(backups_dir)).replace("\\", "/"), "size": p.stat().st_size})

    return templates.TemplateResponse(
        "admin_ops.html",
        {
            "request": request,
            "title": "Admin Ops",
            "logs": logs,
            "backups": backups,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.get("/admin/ops/download/{kind}/{path:path}")
def admin_ops_download(kind: str, path: str, user: UserRow = Depends(require_role("admin"))):
    base = STORAGE_DIR / ("logs" if kind == "logs" else "backups")
    target = (base / path).resolve()
    base_resolved = base.resolve()
    if not str(target).startswith(str(base_resolved)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid path")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Not found")
    return FileResponse(path=target, filename=target.name)


def _pick_existing_dir(primary: Path, fallback: Path) -> Path:
    try:
        if primary.exists():
            # Consider it usable if it has any entries
            for _ in primary.iterdir():
                return primary
    except Exception:
        pass
    return fallback


def _pick_templates_dir(primary: Path, fallback: Path) -> Path:
    # Only pick the new UI templates dir once the essential templates are present.
    try:
        if primary.exists() and (primary / "document_form.html").exists():
            return primary
    except Exception:
        pass
    return fallback


def get_breadcrumbs(path: str):
    breadcrumbs = [{"label": "Trang chủ", "url": "/"}]

    if path.startswith("/contracts"):
        breadcrumbs.append({"label": "Hợp đồng", "url": "/contracts"})
        if "/new" in path:
            breadcrumbs.append({"label": "Tạo mới", "url": None})
    elif path.startswith("/annexes"):
        breadcrumbs.append({"label": "Phụ lục", "url": "/annexes"})
        if "/new" in path:
            breadcrumbs.append({"label": "Tạo mới", "url": None})
    elif path.startswith("/works/import"):
        breadcrumbs.append({"label": "Import tác phẩm", "url": "/works/import"})
    elif path.startswith("/catalogue/upload"):
        breadcrumbs.append({"label": "Upload danh mục", "url": "/catalogue/upload"})
    elif path.startswith("/admin/ops"):
        breadcrumbs.append({"label": "Admin Ops", "url": "/admin/ops"})

    return breadcrumbs


def _pick_static_dir(primary: Path, fallback: Path) -> Path:
    # Only pick the new UI static dir once core assets are present.
    try:
        if primary.exists() and (primary / "css" / "main.css").exists():
            return primary
    except Exception:
        pass
    return fallback


_templates_dir = _pick_templates_dir(UI_TEMPLATES_DIR, WEB_TEMPLATES_DIR)
_static_dir = _pick_static_dir(UI_STATIC_DIR, Path("app/static"))

app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")
templates = Jinja2Templates(directory=str(_templates_dir))

REGION_CODE = "HĐQTGAN-PN"
FIELD_CODE = "MR"
FIELD_NAME = "Sao chép trực tuyến"


def _clean_opt(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def normalize_multi_emails(s: str) -> str:
    raw = _clean_opt(s)
    if not raw:
        return ""
    parts = re.split(r"[;,\s]+", raw)
    parts = [p.strip() for p in parts if p.strip()]
    return ";".join(dict.fromkeys(parts))


def normalize_multi_phones(s: str) -> str:
    raw = _clean_opt(s)
    if not raw:
        return ""
    parts = re.split(r"[;,\s]+", raw)
    parts = [p.strip() for p in parts if p.strip()]
    return ";".join(dict.fromkeys(parts))


def normalize_money_to_int(s: str) -> int:
    raw = _clean_opt(s)
    if not raw:
        return 0
    cleaned = raw.replace("VNĐ", "").replace("VND", "")
    cleaned = cleaned.replace(".", "").replace(",", "")
    cleaned = re.sub(r"\s+", "", cleaned)
    return int(cleaned)


def format_money_number(v: int | None) -> str:
    if v is None:
        return ""
    try:
        return f"{int(v):,}".replace(",", ".")
    except Exception:
        return ""


def format_money_vnd(v: int | None) -> str:
    n = format_money_number(v)
    return f"{n} VNĐ" if n else ""


def normalize_youtube_channel_input(value: str) -> tuple[str, str]:
    s = _clean_opt(value)
    if not s:
        return "", ""
    if s.startswith("http://") or s.startswith("https://"):
        link = s
        cid = _extract_channel_id(s)
        return cid or s, link
    # Could be UC... id
    cid = _extract_channel_id(s) or s
    link = f"https://www.youtube.com/channel/{cid}" if cid.startswith("UC") else ""
    return cid, link


def money_to_vietnamese_words(v: int | None) -> str:
    # Minimal placeholder; existing detailed implementation was removed during refactor.
    if v is None:
        return ""
    try:
        return f"{format_money_number(int(v))} đồng"
    except Exception:
        return ""


def _format_ddmmyyyy(v) -> str:
    if v is None:
        return ""

    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    if not s:
        return ""

    normalized = s.replace("-", "/").replace(".", "/")

    # Accept d/m/yyyy or dd/mm/yyyy (also with '-')
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", normalized)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{d:02d}/{mo:02d}/{y:04d}"

    # If it's not parseable, keep original string (do not hard-fail on import)
    return s


def _extract_channel_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    m = re.search(r"(UC[0-9A-Za-z_-]{10,})", s)
    if m:
        return m.group(1)
    return ""


def _extract_video_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    if s.startswith("="):
        m = re.search(r"watch\?v=([0-9A-Za-z_-]{6,})", s)
        if m:
            return m.group(1)
        m = re.search(r"HYPERLINK\(\"[^\"]*youtu(?:\.be/|be\.com/watch\?v=)([^\"&?#]+)", s, re.IGNORECASE)
        if m:
            return m.group(1)

    m = re.search(r"watch\?v=([0-9A-Za-z_-]{6,})", s)
    if m:
        return m.group(1)
    m = re.search(r"youtu\.be/([0-9A-Za-z_-]{6,})", s)
    if m:
        return m.group(1)

    if re.fullmatch(r"[0-9A-Za-z_-]{6,}", s):
        return s
    return ""


def _normalize_hhmmss(s: str) -> str:
    """Normalize time into hh:mm:ss.

    Accepts:
    - hh:mm:ss
    - mm:ss (will become 00:mm:ss)
    """

    t = (s or "").strip()
    if not t:
        return ""

    parts = [p.strip() for p in t.split(":")]
    if len(parts) == 2:
        hh = 0
        mm = int(parts[0])
        ss = int(parts[1])
    elif len(parts) == 3:
        hh = int(parts[0])
        mm = int(parts[1])
        ss = int(parts[2])
    else:
        raise ValueError("Thời lượng/thời gian phải theo định dạng hh:mm:ss hoặc mm:ss")

    if mm < 0 or mm >= 60 or ss < 0 or ss >= 60 or hh < 0:
        raise ValueError("Thời lượng/thời gian không hợp lệ")

    return f"{hh:02d}:{mm:02d}:{ss:02d}"


def _normalize_time_range(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""

    # accept '-' or '–'
    tt = t.replace("–", "-")
    parts = [p.strip() for p in tt.split("-")]
    if len(parts) != 2:
        raise ValueError("Thời gian phải theo định dạng hh:mm:ss - hh:mm:ss (hoặc mm:ss - mm:ss)")

    start = _normalize_hhmmss(parts[0])
    end = _normalize_hhmmss(parts[1])
    return f"{start} - {end}"


def _parse_import_metadata(ws) -> dict:
    contract_no = ""
    annex_no = ""
    ngay_ky_hop_dong = ""
    ngay_ky_phu_luc = ""
    ten_kenh = ""
    link_kenh = ""

    for r in range(1, min(ws.max_row, 15) + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        s = " ".join(v.split())

        if "HỢP ĐỒNG SỐ" in s.upper() and not contract_no:
            m = re.search(r"HỢP\s*ĐỒNG\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                contract_no = m.group(1).strip()
            m2 = re.search(r"NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_hop_dong = _format_ddmmyyyy(m2.group(1))

        if "PHỤ LỤC SỐ" in s.upper() and not annex_no:
            m = re.search(r"PHỤ\s*LỤC\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                annex_no = m.group(1).strip()
            m2 = re.search(r"PHỤ\s*LỤC\s*SỐ\s*[^\s]+\s*NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_phu_luc = _format_ddmmyyyy(m2.group(1))

        if "YOUTUBE" in s.upper() and not ten_kenh:
            m = re.search(r"YOUTUBE\s+(.+)$", s, re.IGNORECASE)
            if m:
                ten_kenh = m.group(1).strip()

        if ("http://" in s.lower() or "https://" in s.lower()) and not link_kenh:
            m = re.search(r"https?://\S+", s)
            if m:
                link_kenh = m.group(0).strip()

    if not link_kenh:
        for r in range(1, min(ws.max_row, 20) + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and ("http://" in v.lower() or "https://" in v.lower()):
                m = re.search(r"https?://\S+", v)
                if m:
                    link_kenh = m.group(0).strip()
                    break

    return {
        "contract_no": contract_no,
        "annex_no": annex_no,
        "ngay_ky_hop_dong": ngay_ky_hop_dong,
        "ngay_ky_phu_luc": ngay_ky_phu_luc,
        "ten_kenh": ten_kenh,
        "link_kenh": link_kenh,
    }


def _parse_import_table(ws) -> tuple[int, int]:
    header_row = None
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and isinstance(v2, str):
            if v1.strip().upper() == "STT" and "ID" in v2.upper():
                header_row = r
                break
    if not header_row:
        raise ValueError("Không tìm thấy dòng tiêu đề bảng tác phẩm (STT / ID Video)")
    return header_row, header_row + 1


def _norm_header_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())
    return s.strip().lower()


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    # map normalized header -> column index (1-based)
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm_header_cell(ws.cell(row=header_row, column=c).value)
        if not key:
            continue
        # keep first occurrence
        out.setdefault(key, c)
    return out


def _col(hmap: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = _norm_header_cell(n)
        if k in hmap:
            return hmap[k]
    return None


def _year_from_contract_no(contract_no: str) -> int:
    parts = (contract_no or "").split("/")
    if len(parts) >= 2 and parts[1].isdigit():
        return int(parts[1])
    return date.today().year


@app.get("/works/import", response_class=HTMLResponse)
def works_import_form(request: Request, error: str | None = None, message: str | None = None):
    return templates.TemplateResponse(
        "works_import.html",
        {
            "request": request,
            "title": "Import danh sách tác phẩm",
            "error": error,
            "message": message,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/works/import")
async def works_import_submit(
    request: Request,
    import_file: UploadFile = File(...),
    nguoi_thuc_hien: str = Form(""),
    user: UserRow = Depends(require_role("admin", "mod")),
):
    try:
        data = await import_file.read()
        if data is None or len(data) == 0:
            raise ValueError("File import rỗng")
        if len(data) > 25 * 1024 * 1024:
            raise ValueError("File import quá lớn (tối đa 25MB)")

        wb = load_workbook(filename=BytesIO(data), data_only=False)
        ws = wb[wb.sheetnames[0]]

        meta = _parse_import_metadata(ws)
        contract_no = meta.get("contract_no", "")
        annex_no = meta.get("annex_no", "")
        year = _year_from_contract_no(contract_no)

        # Check if uploaded filename matches an existing catalogue file
        uploaded_filename = import_file.filename or ""
        catalogue_dir = STORAGE_EXCEL_DIR / str(year)
        existing_catalogue_path = None
        if uploaded_filename and catalogue_dir.exists():
            potential_path = catalogue_dir / uploaded_filename
            if potential_path.exists():
                existing_catalogue_path = potential_path

        id_channel = _extract_channel_id(meta.get("link_kenh", ""))

        header_row, start_row = _parse_import_table(ws)
        hmap = _build_header_map(ws, header_row)

        c_stt = _col(hmap, "stt") or 1
        c_id_video = _col(hmap, "id video") or 2
        c_code = _col(hmap, "code") or 3
        c_title = _col(hmap, "tên tác phẩm") or 4
        c_author = _col(hmap, "tên tác giả") or 5
        c_composer = _col(hmap, "tên tác giả nhạc") or 6
        c_lyricist = _col(hmap, "tên tác giả lời") or 7
        c_time_range = _col(hmap, "thời gian") or 8
        c_duration = _col(hmap, "thời lượng") or 9
        c_effective = _col(hmap, "ngày bắt đầu") or 10
        c_expiration = _col(hmap, "thời hạn kết thúc") or 11
        c_usage = _col(hmap, "hình thức sử dụng") or 12
        c_rate = _col(hmap, "mức nhuận bút chưa thuế gtgt (vnđ)") or 13
        c_note = _col(hmap, "ghi chú")

        imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        out_rows: list[dict] = []

        for r in range(start_row, ws.max_row + 1):
            stt = ws.cell(row=r, column=c_stt).value
            if stt is None:
                continue

            if isinstance(stt, str) and stt.strip().lower().startswith("cộng"):
                break

            try:
                stt_int = int(str(stt).strip())
            except Exception:
                continue

            id_video = ws.cell(row=r, column=c_id_video).value
            id_link = _extract_video_id(id_video)
            youtube_url = f"https://www.youtube.com/watch?v={id_link}" if id_link else ""

            id_work = ws.cell(row=r, column=c_code).value
            musical_work = ws.cell(row=r, column=c_title).value
            author = ws.cell(row=r, column=c_author).value
            composer = ws.cell(row=r, column=c_composer).value
            lyricist = ws.cell(row=r, column=c_lyricist).value

            time_range_raw = ws.cell(row=r, column=c_time_range).value
            duration_raw = ws.cell(row=r, column=c_duration).value
            effective_date_raw = ws.cell(row=r, column=c_effective).value
            expiration_date_raw = ws.cell(row=r, column=c_expiration).value
            usage_type = ws.cell(row=r, column=c_usage).value
            royalty_rate = ws.cell(row=r, column=c_rate).value
            note = ws.cell(row=r, column=c_note).value if c_note else ""

            time_range = ""
            duration = ""
            if time_range_raw is not None and str(time_range_raw).strip():
                time_range = _normalize_time_range(str(time_range_raw))
            if duration_raw is not None and str(duration_raw).strip():
                duration = _normalize_hhmmss(str(duration_raw))

            effective_date = _format_ddmmyyyy(effective_date_raw)
            expiration_date = _format_ddmmyyyy(expiration_date_raw)

            rr = {
                "year": year,
                "contract_no": contract_no,
                "annex_no": annex_no or "",
                "ngay_ky_hop_dong": meta.get("ngay_ky_hop_dong", ""),
                "ngay_ky_phu_luc": meta.get("ngay_ky_phu_luc", ""),
                "nguoi_thuc_hien": (nguoi_thuc_hien or "").strip(),
                "ten_kenh": meta.get("ten_kenh", ""),
                "id_channel": id_channel,
                "link_kenh": meta.get("link_kenh", ""),
                "stt": stt_int,
                "id_link": id_link,
                "youtube_url": youtube_url,
                "id_work": id_work,
                "musical_work": musical_work,
                "author": author,
                "composer": composer,
                "lyricist": lyricist,
                "time_range": time_range,
                "duration": duration,
                "effective_date": effective_date,
                "expiration_date": expiration_date,
                "usage_type": usage_type,
                "royalty_rate": royalty_rate,
                "note": note,
                "imported_at": imported_at,
            }
            out_rows.append(rr)

        if not contract_no:
            raise ValueError("Không đọc được số hợp đồng trong file import")

        # Insert into DB
        with session_scope() as db:
            objs: list[WorkRow] = []
            for r in out_rows:
                objs.append(
                    WorkRow(
                        year=int(r.get("year") or 0),
                        contract_no=str(r.get("contract_no") or ""),
                        annex_no=(str(r.get("annex_no") or "").strip() or None),
                        ngay_ky_hop_dong=str(r.get("ngay_ky_hop_dong") or ""),
                        ngay_ky_phu_luc=str(r.get("ngay_ky_phu_luc") or ""),
                        nguoi_thuc_hien=str(r.get("nguoi_thuc_hien") or ""),
                        ten_kenh=str(r.get("ten_kenh") or ""),
                        id_channel=str(r.get("id_channel") or ""),
                        link_kenh=str(r.get("link_kenh") or ""),
                        stt=int(r.get("stt")) if r.get("stt") is not None else None,
                        id_link=str(r.get("id_link") or ""),
                        youtube_url=str(r.get("youtube_url") or ""),
                        id_work=str(r.get("id_work") or ""),
                        musical_work=str(r.get("musical_work") or ""),
                        author=str(r.get("author") or ""),
                        composer=str(r.get("composer") or ""),
                        lyricist=str(r.get("lyricist") or ""),
                        time_range=str(r.get("time_range") or ""),
                        duration=str(r.get("duration") or ""),
                        effective_date=str(r.get("effective_date") or ""),
                        expiration_date=str(r.get("expiration_date") or ""),
                        usage_type=str(r.get("usage_type") or ""),
                        royalty_rate=str(r.get("royalty_rate") or ""),
                        note=str(r.get("note") or ""),
                        imported_at=str(r.get("imported_at") or ""),
                    )
                )
            db.bulk_save_objects(objs)

        audit_log(
            log_dir=_LOGS_DIR,
            event={
                "action": "works.import",
                "ip": getattr(getattr(request, "client", None), "host", None),
                "year": year,
                "contract_no": contract_no,
                "annex_no": annex_no or "",
                "rows": len(out_rows),
                "works_table": "works",
            },
        )

        # If uploaded file matches existing catalogue, replace it with the updated version
        if existing_catalogue_path:
            safe_replace_bytes(existing_catalogue_path, data, backup_dir=_BACKUPS_DIR / "files")
            audit_log(
                log_dir=_LOGS_DIR,
                event={
                    "action": "catalogue.replace_from_works_import",
                    "ip": getattr(getattr(request, "client", None), "host", None),
                    "year": year,
                    "contract_no": contract_no,
                    "annex_no": annex_no or "",
                    "file": existing_catalogue_path.name,
                    "bytes": len(data),
                },
            )
            return RedirectResponse(
                url=f"/works/import?message=Đã import {len(out_rows)} dòng vào DB và cập nhật file danh mục {uploaded_filename}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/works/import?message=Đã import {len(out_rows)} dòng vào DB",
            status_code=303,
        )
    except Exception as e:
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(url=f"/works/import?error={msg}", status_code=303)


@app.post("/contracts/{year}/update")
def update_contract(
    request: Request,
    year: int,
    contract_no: str = Form(...),
    ngay_lap_hop_dong: str = Form(...),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form(""),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form("10"),
    user: UserRow = Depends(require_role("admin", "mod")),
):
    try:
        # Calculate money
        pre_vat_value = None
        vat_value = None
        total_value = None
        vat_percent_value = None

        if _clean_opt(so_tien_chua_GTGT):
            pre_vat_value = normalize_money_to_int(_clean_opt(so_tien_chua_GTGT))
            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_value = float(pct_raw.replace(",", "."))
            vat_value = int(round(pre_vat_value * vat_percent_value / 100.0))
            total_value = pre_vat_value + vat_value

        channel_id, channel_link = normalize_youtube_channel_input(kenh_id)

        updated_data = {
            "ngay_lap_hop_dong": date.fromisoformat(ngay_lap_hop_dong),
            "don_vi_ten": _clean_opt(don_vi_ten),
            "don_vi_dia_chi": _clean_opt(don_vi_dia_chi),
            "don_vi_dien_thoai": normalize_multi_phones(don_vi_dien_thoai),
            "don_vi_nguoi_dai_dien": _clean_opt(don_vi_nguoi_dai_dien),
            "don_vi_chuc_vu": _clean_opt(don_vi_chuc_vu),
            "don_vi_mst": _clean_opt(don_vi_mst),
            "don_vi_email": normalize_multi_emails(don_vi_email),
            "kenh_ten": _clean_opt(kenh_ten),
            "kenh_id": channel_id,
            "so_tien_chua_GTGT_value": pre_vat_value,
            "so_tien_chua_GTGT_text": format_money_number(pre_vat_value) if pre_vat_value else "",
            "thue_percent": vat_percent_value,
            "thue_GTGT_value": vat_value,
            "thue_GTGT_text": format_money_number(vat_value) if vat_value else "",
            "so_tien_value": total_value,
            "so_tien_text": format_money_number(total_value) if total_value else "",
            "so_tien_nhuan_but_value": total_value,
            "so_tien_nhuan_but_text": format_money_number(total_value) if total_value else "",
            "so_tien_bang_chu": money_to_vietnamese_words(total_value) if total_value else "",
        }

        success = _db_update_contract_fields(
            year=year,
            contract_no=contract_no,
            annex_no=None,
            updated=updated_data,
        )

        if success:
            audit_log(
                log_dir=_LOGS_DIR,
                event={
                    "action": "contracts.update",
                    "ip": getattr(getattr(request, "client", None), "host", None),
                    "year": year,
                    "contract_no": contract_no,
                    "updated_keys": sorted([k for k in updated_data.keys()]),
                },
            )

        if success:
            return RedirectResponse(url=f"/contracts?year={year}", status_code=303)
        else:
            return RedirectResponse(url=f"/contracts?year={year}&error=Update failed", status_code=303)

    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(url=f"/contracts/{year}/edit?contract_no={quote(contract_no)}&error={str(e)}", status_code=303)


@app.post("/annexes")
def create_annex(
    request: Request,
    contract_no: str = Form(...),
    annex_no: str = Form(""),
    ngay_ky_hop_dong: str = Form(""),
    ngay_ky_phu_luc: str = Form(...),
    linh_vuc: str = Form(""),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form(""),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    so_CCCD: str = Form(""),
    ngay_cap_CCCD: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    nguoi_thuc_hien_email: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form("10"),
):
    try:
        so_phu_luc = annex_no.strip() or None

        # Find contract record first (by year derived from contract_no)
        year = None
        parts = contract_no.split("/")
        if len(parts) >= 2 and parts[1].isdigit():
            year = int(parts[1])
        else:
            year = date.today().year

        # Read from DB for validation + defaults
        contracts = _rows_from_db(year=year)

        # Prevent duplicate annex_no for the same contract
        if so_phu_luc:
            for r in contracts:
                if r.get("contract_no") == contract_no and r.get("annex_no") == so_phu_luc:
                    raise ValueError("Số phụ lục đã tồn tại")

        contract_row: dict | None = None
        for r in contracts:
            if r.get("contract_no") == contract_no:
                contract_row = r
                break

        # Get contract signing date (from form or contract record)
        if ngay_ky_hop_dong and ngay_ky_hop_dong.strip():
            contract_date = date.fromisoformat(ngay_ky_hop_dong)
        elif contract_row and contract_row.get("ngay_lap_hop_dong"):
            contract_date_value = contract_row["ngay_lap_hop_dong"]
            if isinstance(contract_date_value, date):
                if isinstance(contract_date_value, datetime):
                    contract_date = contract_date_value.date()
                else:
                    contract_date = contract_date_value
            elif isinstance(contract_date_value, str):
                contract_date = date.fromisoformat(contract_date_value)
            else:
                return templates.TemplateResponse(
                    "annex_form.html",
                    {
                        "request": request,
                        "title": "Tạo phụ lục",
                        "error": "Không tìm thấy ngày ký hợp đồng",
                        "contracts": contracts,
                        "preview": {},
                        "year": year,
                        "today": date.today().isoformat(),
                        "selected_contract_no": contract_no,
                        "breadcrumbs": get_breadcrumbs(request.url.path),
                    },
                    status_code=400,
                )
        else:
            return templates.TemplateResponse(
                "annex_form.html",
                {
                    "request": request,
                    "title": "Tạo phụ lục",
                    "error": "Vui lòng nhập ngày ký hợp đồng hoặc chọn hợp đồng có sẵn",
                    "contracts": contracts,
                    "preview": {},
                    "year": year,
                    "today": date.today().isoformat(),
                    "selected_contract_no": contract_no,
                    "breadcrumbs": get_breadcrumbs(request.url.path),
                },
                status_code=400,
            )

        contract_date_parts = {
            "ngay_ky_hop_dong": f"{contract_date.day:02d}",
            "thang_ky_hop_dong": f"{contract_date.month:02d}",
            "nam_ky_hop_dong": f"{contract_date.year}",
            "so_hop_dong_day_du": contract_no,
        }

        # Parse annex signing date
        annex_date = date.fromisoformat(ngay_ky_phu_luc)
        annex_date_parts = {
            "ngay_ky_phu_luc": f"{annex_date.day:02d}",
            "thang_ky_phu_luc": f"{annex_date.month:02d}",
            "nam_ky_phu_luc": f"{annex_date.year}",
        }

        # Prefer user overrides; fallback to contract values (if found)
        linh_vuc_value = _clean_opt(linh_vuc) or (contract_row.get("linh_vuc") if contract_row else "") or FIELD_NAME
        don_vi_ten_value = _clean_opt(don_vi_ten) or (contract_row.get("don_vi_ten") if contract_row else "") or ""
        don_vi_dia_chi_value = _clean_opt(don_vi_dia_chi) or (contract_row.get("don_vi_dia_chi") if contract_row else "") or ""
        don_vi_dien_thoai_value = normalize_multi_phones(
            _clean_opt(don_vi_dien_thoai) or (contract_row.get("don_vi_dien_thoai") if contract_row else "") or ""
        )
        don_vi_nguoi_dai_dien_value = _clean_opt(don_vi_nguoi_dai_dien) or (contract_row.get("don_vi_nguoi_dai_dien") if contract_row else "") or ""
        don_vi_chuc_vu_value = _clean_opt(don_vi_chuc_vu) or (contract_row.get("don_vi_chuc_vu") if contract_row else "") or "Giám đốc"
        don_vi_mst_value = _clean_opt(don_vi_mst) or (contract_row.get("don_vi_mst") if contract_row else "") or ""
        don_vi_email_value = normalize_multi_emails(
            _clean_opt(don_vi_email) or (contract_row.get("don_vi_email") if contract_row else "") or ""
        )
        kenh_ten_value = _clean_opt(kenh_ten) or (contract_row.get("kenh_ten") if contract_row else "") or ""

        channel_id_value_raw = _clean_opt(kenh_id) or (contract_row.get("kenh_id") if contract_row else "") or ""
        channel_id_value, channel_link_value = normalize_youtube_channel_input(channel_id_value_raw)

        # Calculate money with VAT
        pre_vat_value = 0
        vat_value = 0
        total_value = 0
        pre_vat_number = ""
        vat_number = ""
        total_number = ""
        total_words = ""

        if _clean_opt(so_tien_chua_GTGT):
            pre_vat_value = normalize_money_to_int(_clean_opt(so_tien_chua_GTGT))
            pre_vat_number = format_money_number(pre_vat_value)

            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_value = float(pct_raw.replace(",", "."))
            if vat_percent_value < 0:
                raise ValueError("Thuế GTGT không hợp lệ")

            vat_value = int(round(pre_vat_value * vat_percent_value / 100.0))
            vat_number = format_money_number(vat_value)

            total_value = pre_vat_value + vat_value
            total_number = format_money_number(total_value)
            total_words = money_to_vietnamese_words(total_value)

        # Build context reusing known placeholders
        context = {
            "contract_no": contract_no,
            "so_hop_dong": contract_no,
            "so_hop_dong_day_du": contract_no,
            "so_phu_luc": so_phu_luc or "",
            "linh_vuc": linh_vuc_value,
            **contract_date_parts,
            **annex_date_parts,
            "don_vi_ten": don_vi_ten_value,
            "don_vi_dia_chi": don_vi_dia_chi_value,
            "don_vi_dien_thoai": don_vi_dien_thoai_value,
            "don_vi_nguoi_dai_dien": don_vi_nguoi_dai_dien_value,
            "don_vi_chuc_vu": don_vi_chuc_vu_value,
            "don_vi_mst": don_vi_mst_value,
            "don_vi_email": don_vi_email_value,
            "so_CCCD": so_CCCD or "",
            "ngay_cap_CCCD": ngay_cap_CCCD or "",
            "kenh_ten": kenh_ten_value,
            "kenh_id": channel_id_value,
            "nguoi_thuc_hien_email": nguoi_thuc_hien_email or "",
            "so_tien_nhuan_but": total_number,
            "so_tien_chua_GTGT": pre_vat_number,
            "thue_GTGT": vat_number,
            "so_tien_GTGT": total_number,
            "so_tien": total_number,
            "so_tien_bang_chu": total_words,
            "thue_percent": str(int(vat_percent_value)) if vat_percent_value else "10",
            "TEN_DON_VI": don_vi_ten_value,
            "ten_don_vi": don_vi_ten_value,
            "dia_chi": don_vi_dia_chi_value,
            "so_dien_thoai": don_vi_dien_thoai_value,
            "NGUOI_DAI_DIEN": don_vi_nguoi_dai_dien_value,
            "nguoi_dai_dien": don_vi_nguoi_dai_dien_value,
            "CHUC_VU": don_vi_chuc_vu_value,
            "chuc_vu": don_vi_chuc_vu_value,
            "ma_so_thue": don_vi_mst_value,
            "email": don_vi_email_value,
            "ten_kenh": kenh_ten_value,
            "link_kenh": channel_link_value,
            "so_tien_nhuan_but": total_number,
            "so_tien_chua_GTGT": pre_vat_number,
            "thue_GTGT": vat_number,
            "so_tien_GTGT": total_number,
            "so_tien": total_number,
            "so_tien_bang_chu": total_words,
            "thue_percent": str(int(vat_percent_value)) if vat_percent_value else "10",
        }

        # Save DOCX
        out_docx_dir = STORAGE_DOCX_DIR / str(year)
        out_docx_dir.mkdir(parents=True, exist_ok=True)
        filename = build_docx_filename(
            year=year,
            so_hop_dong_4=_parse_so_hop_dong_4(contract_no),
            so_phu_luc=so_phu_luc,
            linh_vuc=linh_vuc_value,
            kenh_ten=kenh_ten_value,
        )
        out_docx_path = out_docx_dir / filename
        if out_docx_path.exists():
            stem = out_docx_path.stem
            out_docx_path = out_docx_dir / f"{stem}_{date.today().strftime('%Y%m%d')}.docx"

        render_contract_docx(template_path=ANNEX_TEMPLATE_PATH, output_path=out_docx_path, context=context)

        # Export annex catalogue Excel from template
        out_excel_dir = STORAGE_EXCEL_DIR / str(year)
        out_excel_dir.mkdir(parents=True, exist_ok=True)
        catalogue_name = out_docx_path.with_suffix(".xlsx").name
        out_catalogue_path = out_excel_dir / catalogue_name

        catalogue_context = dict(context)
        catalogue_context["so_hop_dong_day_du"] = contract_no
        catalogue_context["ngay_ky_hop_dong"] = contract_date.strftime("%d/%m/%Y")
        catalogue_context["ngay_ky_phu_luc"] = annex_date.strftime("%d/%m/%Y")
        export_catalogue_excel(
            template_path=ANNEX_CATALOGUE_TEMPLATE_PATH,
            output_path=out_catalogue_path,
            context=catalogue_context,
            sheet_name="Final",
        )

        # Determine VAT percent value
        vat_percent_final = None
        if _clean_opt(so_tien_chua_GTGT):
            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_final = float(pct_raw.replace(",", "."))

        annex_record = ContractRecord(
            contract_no=contract_no,
            contract_year=year,
            annex_no=so_phu_luc,
            ngay_lap_hop_dong=contract_date,
            linh_vuc=linh_vuc_value,
            region_code=REGION_CODE,
            field_code=FIELD_CODE,
            don_vi_ten=don_vi_ten_value,
            don_vi_dia_chi=don_vi_dia_chi_value,
            don_vi_dien_thoai=don_vi_dien_thoai_value,
            don_vi_nguoi_dai_dien=don_vi_nguoi_dai_dien_value,
            don_vi_chuc_vu=don_vi_chuc_vu_value,
            don_vi_mst=don_vi_mst_value,
            don_vi_email=don_vi_email_value,
            kenh_ten=kenh_ten_value,
            kenh_id=channel_id_value,
            so_tien_nhuan_but_value=total_value if total_value else None,
            so_tien_nhuan_but_text=format_money_vnd(total_value) if total_value else None,
            so_tien_chua_GTGT_value=pre_vat_value if pre_vat_value else None,
            so_tien_chua_GTGT_text=format_money_vnd(pre_vat_value) if pre_vat_value else None,
            thue_percent=vat_percent_final,
            thue_GTGT_value=vat_value if vat_value else None,
            thue_GTGT_text=format_money_vnd(vat_value) if vat_value else None,
            so_tien_value=total_value if total_value else None,
            so_tien_text=format_money_vnd(total_value) if total_value else None,
            so_tien_bang_chu=total_words if total_words else None,
            docx_path=str(out_docx_path),
            catalogue_path=str(out_catalogue_path),
        )
        _db_upsert_contract_record(
            record={
                "contract_no": contract_no,
                "contract_year": year,
                "annex_no": so_phu_luc,
                "ngay_lap_hop_dong": contract_date,
                "linh_vuc": linh_vuc_value,
                "region_code": REGION_CODE,
                "field_code": FIELD_CODE,
                "don_vi_ten": don_vi_ten_value,
                "don_vi_dia_chi": don_vi_dia_chi_value,
                "don_vi_dien_thoai": don_vi_dien_thoai_value,
                "don_vi_nguoi_dai_dien": don_vi_nguoi_dai_dien_value,
                "don_vi_chuc_vu": don_vi_chuc_vu_value,
                "don_vi_mst": don_vi_mst_value,
                "don_vi_email": don_vi_email_value,
                "so_CCCD": _clean_opt(so_CCCD),
                "ngay_cap_CCCD": _clean_opt(ngay_cap_CCCD),
                "kenh_ten": kenh_ten_value,
                "kenh_id": channel_id_value,
                "nguoi_thuc_hien_email": normalize_multi_emails(nguoi_thuc_hien_email),
                "so_tien_nhuan_but_value": total_value if total_value else None,
                "so_tien_nhuan_but_text": format_money_number(total_value) if total_value else None,
                "so_tien_chua_GTGT_value": pre_vat_value if pre_vat_value else None,
                "so_tien_chua_GTGT_text": format_money_number(pre_vat_value) if pre_vat_value else None,
                "thue_percent": vat_percent_value if vat_percent_value else None,
                "thue_GTGT_value": vat_value if vat_value else None,
                "thue_GTGT_text": format_money_number(vat_value) if vat_value else None,
                "so_tien_value": total_value if total_value else None,
                "so_tien_text": format_money_number(total_value) if total_value else None,
                "so_tien_bang_chu": total_words if total_words else None,
                "docx_path": str(out_docx_path),
                "catalogue_path": str(out_catalogue_path),
            }
        )

        audit_log(
            log_dir=_LOGS_DIR,
            event={
                "action": "annexes.create",
                "ip": getattr(getattr(request, "client", None), "host", None),
                "year": year,
                "contract_no": contract_no,
                "annex_no": so_phu_luc,
                "docx": out_docx_path.name,
                "catalogue": out_catalogue_path.name,
            },
        )

        return RedirectResponse(
            url=(
                f"/contracts?year={year}"
                f"&download=/download/{year}/{out_docx_path.name}"
                f"&download2=/download_excel/{year}/{out_catalogue_path.name}"
            ),
            status_code=303,
        )
    except Exception as e:
        return RedirectResponse(url=f"/documents/new?doc_type=annex&error={str(e)}", status_code=303)


@app.get("/documents/new", response_class=HTMLResponse)
def document_form_unified(
    request: Request,
    doc_type: str | None = None,
    year: int | None = None,
    contract_no: str | None = None,
    error: str | None = None,
):
    y = year or date.today().year
    contracts = _rows_from_db(year=y)

    # Filter: only show base contracts (no annex_no) for dropdown
    contracts = [r for r in contracts if not r.get("annex_no")]

    preview: dict = {}
    if contract_no and doc_type == "annex":
        for r in contracts:
            if r.get("contract_no") == contract_no:
                preview = r.copy()
                if "ngay_lap_hop_dong" in preview:
                    val = preview["ngay_lap_hop_dong"]
                    if isinstance(val, date):
                        if isinstance(val, datetime):
                            val = val.date()
                        preview["ngay_lap_hop_dong"] = val.isoformat()
                break

    return templates.TemplateResponse(
        "document_form.html",
        {
            "request": request,
            "title": "Tạo tài liệu",
            "error": error,
            "doc_type": doc_type or "contract",
            "contracts": contracts,
            "preview": preview,
            "year": y,
            "today": date.today().isoformat(),
            "selected_contract_no": contract_no or "",
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/documents")
def create_document_unified(
    request: Request,
    doc_type: str = Form(...),
    ngay_lap_hop_dong: str = Form(""),
    so_hop_dong_4: str = Form(""),
    contract_no: str = Form(""),
    annex_no: str = Form(""),
    ngay_ky_hop_dong: str = Form(""),
    ngay_ky_phu_luc: str = Form(""),
    linh_vuc: str = Form(""),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form(""),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    so_CCCD: str = Form(""),
    ngay_cap_CCCD: str = Form(""),
    nguoi_thuc_hien_email: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form("10"),
):
    if doc_type == "contract":
        # Redirect to contract creation
        return create_contract(
            request=request,
            ngay_lap_hop_dong=ngay_lap_hop_dong,
            so_hop_dong_4=so_hop_dong_4,
            linh_vuc=linh_vuc or FIELD_NAME,
            don_vi_ten=don_vi_ten,
            don_vi_dia_chi=don_vi_dia_chi,
            don_vi_dien_thoai=don_vi_dien_thoai,
            don_vi_nguoi_dai_dien=don_vi_nguoi_dai_dien,
            don_vi_chuc_vu=don_vi_chuc_vu,
            don_vi_mst=don_vi_mst,
            don_vi_email=don_vi_email,
            so_CCCD=so_CCCD,
            ngay_cap_CCCD=ngay_cap_CCCD,
            nguoi_thuc_hien_email=nguoi_thuc_hien_email,
            kenh_ten=kenh_ten,
            kenh_id=kenh_id,
            so_tien_chua_GTGT=so_tien_chua_GTGT,
            thue_percent=thue_percent,
        )
    elif doc_type == "annex":
        # Redirect to annex creation
        return create_annex(
            request=request,
            contract_no=contract_no,
            annex_no=annex_no,
            ngay_ky_hop_dong=ngay_ky_hop_dong,
            ngay_ky_phu_luc=ngay_ky_phu_luc,
            linh_vuc=linh_vuc,
            don_vi_ten=don_vi_ten,
            don_vi_dia_chi=don_vi_dia_chi,
            don_vi_dien_thoai=don_vi_dien_thoai,
            don_vi_nguoi_dai_dien=don_vi_nguoi_dai_dien,
            don_vi_chuc_vu=don_vi_chuc_vu,
            don_vi_mst=don_vi_mst,
            don_vi_email=don_vi_email,
            so_CCCD=so_CCCD,
            ngay_cap_CCCD=ngay_cap_CCCD,
            kenh_ten=kenh_ten,
            kenh_id=kenh_id,
            nguoi_thuc_hien_email=nguoi_thuc_hien_email,
            so_tien_chua_GTGT=so_tien_chua_GTGT,
            thue_percent=thue_percent,
        )
    else:
        return RedirectResponse(url="/documents/new?error=Invalid document type", status_code=303)


@app.get("/storage/excel/download/{year}")
def download_contracts_excel(year: int):
    if not _db_available():
        return JSONResponse({"error": "DB không tồn tại"}, status_code=500)

    data = _export_contracts_excel_bytes(year=year)
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="contracts_{year}.xlsx"'},
    )


@app.get("/storage/excel/works/download/{year}")
def download_works_excel(year: int):
    if not _db_available():
        return JSONResponse({"error": "DB không tồn tại"}, status_code=500)

    data = _export_works_excel_bytes(year=year)
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="works_contract_{year}.xlsx"'},
    )


@app.get("/storage/files/{year}")
def list_saved_files(year: int):
    year_dir_docx = STORAGE_DOCX_DIR / str(year)
    year_dir_excel = STORAGE_EXCEL_DIR / str(year)

    files = []

    if year_dir_docx.exists():
        for f in year_dir_docx.glob("*.docx"):
            files.append({
                "name": f.name,
                "type": "docx",
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "url": f"/storage/docx/{year}/{f.name}"
            })

    if year_dir_excel.exists():
        for f in year_dir_excel.glob("*.xlsx"):
            files.append({
                "name": f.name,
                "type": "xlsx",
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "url": f"/storage/excel/{year}/{f.name}"
            })

    files.sort(key=lambda x: x["modified"], reverse=True)
    return JSONResponse(files)


@app.get("/storage/docx/{year}/{filename}")
def download_docx_file(year: int, filename: str):
    file_path = STORAGE_DOCX_DIR / str(year) / filename
    if not file_path.exists():
        return JSONResponse({"error": "File không tồn tại"}, status_code=404)
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )


@app.get("/storage/excel/{year}/{filename}")
def download_excel_file(year: int, filename: str):
    file_path = STORAGE_EXCEL_DIR / str(year) / filename
    if not file_path.exists():
        return JSONResponse({"error": "File không tồn tại"}, status_code=404)
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
