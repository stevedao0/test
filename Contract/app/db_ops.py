from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app.db import DB_PATH, session_scope
from app.db_models import ContractRecordRow, WorkRow
from app.services.excel_store import HEADERS, WORKS_HEADERS


_HEADER_FONT = Font(bold=True)


def _db_available() -> bool:
    try:
        return DB_PATH.exists()
    except Exception:
        return False


def _pick_latest_contract_year(default_year: int) -> int:
    if not _db_available():
        return default_year

    try:
        with session_scope() as db:
            row = db.query(ContractRecordRow.contract_year).order_by(ContractRecordRow.contract_year.desc()).first()
            if row and row[0]:
                return int(row[0])
    except Exception:
        return default_year

    return default_year


def _rows_from_db(*, year: int) -> list[dict]:
    if not _db_available():
        return []

    with session_scope() as db:
        q = db.query(ContractRecordRow).filter(ContractRecordRow.contract_year == year)
        out: list[dict] = []
        for r in q.all():
            out.append(
                {
                    "contract_no": r.contract_no,
                    "contract_year": r.contract_year,
                    "annex_no": r.annex_no,
                    "ngay_lap_hop_dong": r.ngay_lap_hop_dong,
                    "linh_vuc": r.linh_vuc,
                    "region_code": r.region_code,
                    "field_code": r.field_code,
                    "don_vi_ten": r.don_vi_ten,
                    "don_vi_dia_chi": r.don_vi_dia_chi,
                    "don_vi_dien_thoai": r.don_vi_dien_thoai,
                    "don_vi_nguoi_dai_dien": r.don_vi_nguoi_dai_dien,
                    "don_vi_chuc_vu": r.don_vi_chuc_vu,
                    "don_vi_mst": r.don_vi_mst,
                    "don_vi_email": r.don_vi_email,
                    "so_CCCD": r.so_cccd,
                    "ngay_cap_CCCD": r.ngay_cap_cccd,
                    "kenh_ten": r.kenh_ten,
                    "kenh_id": r.kenh_id,
                    "nguoi_thuc_hien_email": r.nguoi_thuc_hien_email,
                    "so_tien_nhuan_but_value": r.so_tien_nhuan_but_value,
                    "so_tien_nhuan_but_text": r.so_tien_nhuan_but_text,
                    "so_tien_chua_GTGT_value": r.so_tien_chua_gtgt_value,
                    "so_tien_chua_GTGT_text": r.so_tien_chua_gtgt_text,
                    "thue_percent": r.thue_percent,
                    "thue_GTGT_value": r.thue_gtgt_value,
                    "thue_GTGT_text": r.thue_gtgt_text,
                    "so_tien_value": r.so_tien_value,
                    "so_tien_text": r.so_tien_text,
                    "so_tien_bang_chu": r.so_tien_bang_chu,
                    "docx_path": r.docx_path,
                    "catalogue_path": r.catalogue_path,
                }
            )
        return out


def _db_get_contract_row(*, year: int, contract_no: str, annex_no: str | None) -> ContractRecordRow | None:
    with session_scope() as db:
        return (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
            .filter(
                ContractRecordRow.annex_no.is_(annex_no)
                if annex_no is None
                else (ContractRecordRow.annex_no == annex_no)
            )
            .first()
        )


def _db_upsert_contract_record(*, record: dict) -> None:
    year = int(record.get("contract_year") or 0)
    contract_no = str(record.get("contract_no") or "")
    annex_no = record.get("annex_no")
    annex_no = (str(annex_no).strip() if annex_no is not None else None) or None

    with session_scope() as db:
        q = (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
        )
        if annex_no is None:
            q = q.filter(ContractRecordRow.annex_no.is_(None))
        else:
            q = q.filter(ContractRecordRow.annex_no == annex_no)

        row = q.first()
        if row is None:
            row = ContractRecordRow(contract_year=year, contract_no=contract_no, annex_no=annex_no)
            db.add(row)

        for k, v in record.items():
            if k in ("so_CCCD", "ngay_cap_CCCD"):
                continue
            if hasattr(row, k):
                setattr(row, k, v)

        if "so_CCCD" in record:
            row.so_cccd = record.get("so_CCCD")
        if "ngay_cap_CCCD" in record:
            row.ngay_cap_cccd = record.get("ngay_cap_CCCD")


def _db_update_contract_fields(*, year: int, contract_no: str, annex_no: str | None, updated: dict) -> bool:
    annex_no = (annex_no.strip() if isinstance(annex_no, str) else annex_no) or None

    with session_scope() as db:
        q = (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
        )
        if annex_no is None:
            q = q.filter(ContractRecordRow.annex_no.is_(None))
        else:
            q = q.filter(ContractRecordRow.annex_no == annex_no)

        row = q.first()
        if row is None:
            return False

        for k, v in updated.items():
            if k in ("so_CCCD", "ngay_cap_CCCD"):
                continue
            if hasattr(row, k):
                setattr(row, k, v)

        if "so_CCCD" in updated:
            row.so_cccd = updated.get("so_CCCD")
        if "ngay_cap_CCCD" in updated:
            row.ngay_cap_cccd = updated.get("ngay_cap_CCCD")

        return True


def _db_delete_contract_record(*, year: int, contract_no: str, annex_no: str | None) -> bool:
    annex_no = (annex_no.strip() if isinstance(annex_no, str) else annex_no) or None
    with session_scope() as db:
        q = (
            db.query(ContractRecordRow)
            .filter(ContractRecordRow.contract_year == year)
            .filter(ContractRecordRow.contract_no == contract_no)
        )
        if annex_no is None:
            q = q.filter(ContractRecordRow.annex_no.is_(None))
        else:
            q = q.filter(ContractRecordRow.annex_no == annex_no)

        row = q.first()
        if row is None:
            return False
        db.delete(row)
        return True


def _xlsx_bytes_from_rows(*, sheet_name: str, headers: list[str], rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT

    for r_idx, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c, value=row.get(h))

    bio = BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue()


def _export_contracts_excel_bytes(*, year: int) -> bytes:
    with session_scope() as db:
        q = db.query(ContractRecordRow).filter(ContractRecordRow.contract_year == year)
        db_rows = q.all()

    rows: list[dict] = []
    for r in db_rows:
        rows.append(
            {
                "contract_no": r.contract_no,
                "contract_year": r.contract_year,
                "annex_no": r.annex_no,
                "ngay_lap_hop_dong": r.ngay_lap_hop_dong,
                "linh_vuc": r.linh_vuc,
                "region_code": r.region_code,
                "field_code": r.field_code,
                "don_vi_ten": r.don_vi_ten,
                "don_vi_dia_chi": r.don_vi_dia_chi,
                "don_vi_dien_thoai": r.don_vi_dien_thoai,
                "don_vi_nguoi_dai_dien": r.don_vi_nguoi_dai_dien,
                "don_vi_chuc_vu": r.don_vi_chuc_vu,
                "don_vi_mst": r.don_vi_mst,
                "don_vi_email": r.don_vi_email,
                "so_CCCD": r.so_cccd,
                "ngay_cap_CCCD": r.ngay_cap_cccd,
                "kenh_ten": r.kenh_ten,
                "kenh_id": r.kenh_id,
                "nguoi_thuc_hien_email": r.nguoi_thuc_hien_email,
                "so_tien_nhuan_but_value": r.so_tien_nhuan_but_value,
                "so_tien_nhuan_but_text": r.so_tien_nhuan_but_text,
                "so_tien_chua_GTGT_value": r.so_tien_chua_gtgt_value,
                "so_tien_chua_GTGT_text": r.so_tien_chua_gtgt_text,
                "thue_percent": r.thue_percent,
                "thue_GTGT_value": r.thue_gtgt_value,
                "thue_GTGT_text": r.thue_gtgt_text,
                "so_tien_value": r.so_tien_value,
                "so_tien_text": r.so_tien_text,
                "so_tien_bang_chu": r.so_tien_bang_chu,
                "docx_path": r.docx_path,
                "catalogue_path": r.catalogue_path,
            }
        )

    return _xlsx_bytes_from_rows(sheet_name="Contracts", headers=list(HEADERS), rows=rows)


def _export_works_excel_bytes(*, year: int) -> bytes:
    with session_scope() as db:
        q = db.query(WorkRow).filter(WorkRow.year == year)
        db_rows = q.all()

    rows: list[dict] = []
    for r in db_rows:
        rows.append(
            {
                "year": r.year,
                "contract_no": r.contract_no,
                "annex_no": r.annex_no,
                "ngay_ky_hop_dong": r.ngay_ky_hop_dong,
                "ngay_ky_phu_luc": r.ngay_ky_phu_luc,
                "nguoi_thuc_hien": r.nguoi_thuc_hien,
                "ten_kenh": r.ten_kenh,
                "id_channel": r.id_channel,
                "link_kenh": r.link_kenh,
                "stt": r.stt,
                "id_link": r.id_link,
                "youtube_url": r.youtube_url,
                "id_work": r.id_work,
                "musical_work": r.musical_work,
                "author": r.author,
                "composer": r.composer,
                "lyricist": r.lyricist,
                "time_range": r.time_range,
                "duration": r.duration,
                "effective_date": r.effective_date,
                "expiration_date": r.expiration_date,
                "usage_type": r.usage_type,
                "royalty_rate": r.royalty_rate,
                "note": r.note,
                "imported_at": r.imported_at,
            }
        )

    return _xlsx_bytes_from_rows(sheet_name="Works", headers=list(WORKS_HEADERS), rows=rows)


def _maybe_backup_file(path: Path, *, backup_dir: Path) -> None:
    try:
        if path.exists() and path.is_file():
            backup_dir.mkdir(parents=True, exist_ok=True)
            target = backup_dir / path.name
            if target.exists():
                target = backup_dir / f"{target.stem}_dup{target.suffix}"
            path.replace(target)
    except Exception:
        pass
