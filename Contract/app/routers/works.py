from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook

from app.auth import require_role
from app.config import STORAGE_DIR, STORAGE_EXCEL_DIR
from app.db import session_scope
from app.db_models import UserRow, WorkRow
from app.services.safety import audit_log, safe_replace_bytes
from app.utils import extract_channel_id, get_breadcrumbs


router = APIRouter()


@router.get("/works/import", response_class=HTMLResponse)
def works_import_form(request: Request, error: str | None = None, message: str | None = None):
    templates = request.app.state.templates
    return templates.TemplateResponse(
        "works_import.html",
        {
            "request": request,
            "title": "Import danh sách tác phẩm",
            "error": error,
            "message": message,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@router.post("/works/import")
async def works_import_submit(
    request: Request,
    import_file: UploadFile = File(...),
    nguoi_thuc_hien: str = Form(""),
    user: UserRow = Depends(require_role("admin", "mod")),
):
    backups_dir = STORAGE_DIR / "backups"
    logs_dir = STORAGE_DIR / "logs"

    try:
        data = await import_file.read()
        if data is None or len(data) == 0:
            raise ValueError("File import rỗng")
        if len(data) > 25 * 1024 * 1024:
            raise ValueError("File import quá lớn (tối đa 25MB)")

        wb = load_workbook(filename=BytesIO(data), data_only=False)
        ws = wb[wb.sheetnames[0]]

        meta = _parse_import_metadata(ws)
        contract_no = meta.get("contract_no", "")
        annex_no = meta.get("annex_no", "")
        year = _year_from_contract_no(contract_no)

        uploaded_filename = import_file.filename or ""
        catalogue_dir = STORAGE_EXCEL_DIR / str(year)
        existing_catalogue_path = None
        if uploaded_filename and catalogue_dir.exists():
            potential_path = catalogue_dir / uploaded_filename
            if potential_path.exists():
                existing_catalogue_path = potential_path

        id_channel = extract_channel_id(meta.get("link_kenh", ""))

        header_row, start_row = _parse_import_table(ws)
        hmap = _build_header_map(ws, header_row)

        c_stt = _col(hmap, "stt") or 1
        c_id_video = _col(hmap, "id video") or 2
        c_code = _col(hmap, "code") or 3
        c_title = _col(hmap, "tên tác phẩm") or 4
        c_author = _col(hmap, "tên tác giả") or 5
        c_composer = _col(hmap, "tên tác giả nhạc") or 6
        c_lyricist = _col(hmap, "tên tác giả lời") or 7
        c_time_range = _col(hmap, "thời gian") or 8
        c_duration = _col(hmap, "thời lượng") or 9
        c_effective = _col(hmap, "ngày bắt đầu") or 10
        c_expiration = _col(hmap, "thời hạn kết thúc") or 11
        c_usage = _col(hmap, "hình thức sử dụng") or 12
        c_rate = _col(hmap, "mức nhuận bút chưa thuế gtgt (vnđ)") or 13
        c_note = _col(hmap, "ghi chú")

        imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        out_rows: list[dict] = []

        for r in range(start_row, ws.max_row + 1):
            stt = ws.cell(row=r, column=c_stt).value
            if stt is None:
                continue

            if isinstance(stt, str) and stt.strip().lower().startswith("cộng"):
                break

            try:
                stt_int = int(str(stt).strip())
            except Exception:
                continue

            id_video = ws.cell(row=r, column=c_id_video).value
            id_link = _extract_video_id(id_video)
            youtube_url = f"https://www.youtube.com/watch?v={id_link}" if id_link else ""

            id_work = ws.cell(row=r, column=c_code).value
            musical_work = ws.cell(row=r, column=c_title).value
            author = ws.cell(row=r, column=c_author).value
            composer = ws.cell(row=r, column=c_composer).value
            lyricist = ws.cell(row=r, column=c_lyricist).value

            time_range_raw = ws.cell(row=r, column=c_time_range).value
            duration_raw = ws.cell(row=r, column=c_duration).value
            effective_date_raw = ws.cell(row=r, column=c_effective).value
            expiration_date_raw = ws.cell(row=r, column=c_expiration).value
            usage_type = ws.cell(row=r, column=c_usage).value
            royalty_rate = ws.cell(row=r, column=c_rate).value
            note = ws.cell(row=r, column=c_note).value if c_note else ""

            time_range = ""
            duration = ""
            if time_range_raw is not None and str(time_range_raw).strip():
                time_range = _normalize_time_range(str(time_range_raw))
            if duration_raw is not None and str(duration_raw).strip():
                duration = _normalize_hhmmss(str(duration_raw))

            effective_date = _format_ddmmyyyy(effective_date_raw)
            expiration_date = _format_ddmmyyyy(expiration_date_raw)

            rr = {
                "year": year,
                "contract_no": contract_no,
                "annex_no": annex_no or "",
                "ngay_ky_hop_dong": meta.get("ngay_ky_hop_dong", ""),
                "ngay_ky_phu_luc": meta.get("ngay_ky_phu_luc", ""),
                "nguoi_thuc_hien": (nguoi_thuc_hien or "").strip(),
                "ten_kenh": meta.get("ten_kenh", ""),
                "id_channel": id_channel,
                "link_kenh": meta.get("link_kenh", ""),
                "stt": stt_int,
                "id_link": id_link,
                "youtube_url": youtube_url,
                "id_work": id_work,
                "musical_work": musical_work,
                "author": author,
                "composer": composer,
                "lyricist": lyricist,
                "time_range": time_range,
                "duration": duration,
                "effective_date": effective_date,
                "expiration_date": expiration_date,
                "usage_type": usage_type,
                "royalty_rate": royalty_rate,
                "note": note,
                "imported_at": imported_at,
            }
            out_rows.append(rr)

        if not contract_no:
            raise ValueError("Không đọc được số hợp đồng trong file import")

        with session_scope() as db:
            objs: list[WorkRow] = []
            for rr in out_rows:
                objs.append(
                    WorkRow(
                        year=int(rr.get("year") or 0),
                        contract_no=str(rr.get("contract_no") or ""),
                        annex_no=(str(rr.get("annex_no") or "").strip() or None),
                        ngay_ky_hop_dong=str(rr.get("ngay_ky_hop_dong") or ""),
                        ngay_ky_phu_luc=str(rr.get("ngay_ky_phu_luc") or ""),
                        nguoi_thuc_hien=str(rr.get("nguoi_thuc_hien") or ""),
                        ten_kenh=str(rr.get("ten_kenh") or ""),
                        id_channel=str(rr.get("id_channel") or ""),
                        link_kenh=str(rr.get("link_kenh") or ""),
                        stt=int(rr.get("stt")) if rr.get("stt") is not None else None,
                        id_link=str(rr.get("id_link") or ""),
                        youtube_url=str(rr.get("youtube_url") or ""),
                        id_work=str(rr.get("id_work") or ""),
                        musical_work=str(rr.get("musical_work") or ""),
                        author=str(rr.get("author") or ""),
                        composer=str(rr.get("composer") or ""),
                        lyricist=str(rr.get("lyricist") or ""),
                        time_range=str(rr.get("time_range") or ""),
                        duration=str(rr.get("duration") or ""),
                        effective_date=str(rr.get("effective_date") or ""),
                        expiration_date=str(rr.get("expiration_date") or ""),
                        usage_type=str(rr.get("usage_type") or ""),
                        royalty_rate=str(rr.get("royalty_rate") or ""),
                        note=str(rr.get("note") or ""),
                        imported_at=str(rr.get("imported_at") or ""),
                    )
                )
            db.bulk_save_objects(objs)

        audit_log(
            log_dir=logs_dir,
            event={
                "action": "works.import",
                "ip": getattr(getattr(request, "client", None), "host", None),
                "year": year,
                "contract_no": contract_no,
                "annex_no": annex_no or "",
                "rows": len(out_rows),
                "works_table": "works",
                "actor": user.username,
            },
        )

        if existing_catalogue_path:
            safe_replace_bytes(existing_catalogue_path, data, backup_dir=backups_dir / "files")
            audit_log(
                log_dir=logs_dir,
                event={
                    "action": "catalogue.replace_from_works_import",
                    "ip": getattr(getattr(request, "client", None), "host", None),
                    "year": year,
                    "contract_no": contract_no,
                    "annex_no": annex_no or "",
                    "file": existing_catalogue_path.name,
                    "bytes": len(data),
                    "actor": user.username,
                },
            )
            return RedirectResponse(
                url=f"/works/import?message=Đã import {len(out_rows)} dòng vào DB và cập nhật file danh mục {uploaded_filename}",
                status_code=303,
            )

        return RedirectResponse(url=f"/works/import?message=Đã import {len(out_rows)} dòng vào DB", status_code=303)

    except Exception as e:
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(url=f"/works/import?error={msg}", status_code=303)


def _format_ddmmyyyy(v) -> str:
    if v is None:
        return ""

    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    if not s:
        return ""

    normalized = s.replace("-", "/").replace(".", "/")
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", normalized)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{d:02d}/{mo:02d}/{y:04d}"

    return s


def _extract_video_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    if s.startswith("="):
        m = re.search(r"watch\?v=([0-9A-Za-z_-]{6,})", s)
        if m:
            return m.group(1)
        m = re.search(r"HYPERLINK\(\"[^\"]*youtu(?:\.be/|be\.com/watch\?v=)([^\"&?#]+)", s, re.IGNORECASE)
        if m:
            return m.group(1)

    m = re.search(r"watch\?v=([0-9A-Za-z_-]{6,})", s)
    if m:
        return m.group(1)
    m = re.search(r"youtu\.be/([0-9A-Za-z_-]{6,})", s)
    if m:
        return m.group(1)

    if re.fullmatch(r"[0-9A-Za-z_-]{6,}", s):
        return s
    return ""


def _normalize_hhmmss(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""

    parts = [p.strip() for p in t.split(":")]
    if len(parts) == 2:
        hh = 0
        mm = int(parts[0])
        ss = int(parts[1])
    elif len(parts) == 3:
        hh = int(parts[0])
        mm = int(parts[1])
        ss = int(parts[2])
    else:
        raise ValueError("Thời lượng/thời gian phải theo định dạng hh:mm:ss hoặc mm:ss")

    if mm < 0 or mm >= 60 or ss < 0 or ss >= 60 or hh < 0:
        raise ValueError("Thời lượng/thời gian không hợp lệ")

    return f"{hh:02d}:{mm:02d}:{ss:02d}"


def _normalize_time_range(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""

    tt = t.replace("–", "-")
    parts = [p.strip() for p in tt.split("-")]
    if len(parts) != 2:
        raise ValueError("Thời gian phải theo định dạng hh:mm:ss - hh:mm:ss (hoặc mm:ss - mm:ss)")

    start = _normalize_hhmmss(parts[0])
    end = _normalize_hhmmss(parts[1])
    return f"{start} - {end}"


def _parse_import_metadata(ws) -> dict:
    contract_no = ""
    annex_no = ""
    ngay_ky_hop_dong = ""
    ngay_ky_phu_luc = ""
    ten_kenh = ""
    link_kenh = ""

    for r in range(1, min(ws.max_row, 15) + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        s = " ".join(v.split())

        if "HỢP ĐỒNG SỐ" in s.upper() and not contract_no:
            m = re.search(r"HỢP\s*ĐỒNG\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                contract_no = m.group(1).strip()
            m2 = re.search(r"NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_hop_dong = _format_ddmmyyyy(m2.group(1))

        if "PHỤ LỤC SỐ" in s.upper() and not annex_no:
            m = re.search(r"PHỤ\s*LỤC\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                annex_no = m.group(1).strip()
            m2 = re.search(r"PHỤ\s*LỤC\s*SỐ\s*[^\s]+\s*NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_phu_luc = _format_ddmmyyyy(m2.group(1))

        if "YOUTUBE" in s.upper() and not ten_kenh:
            m = re.search(r"YOUTUBE\s+(.+)$", s, re.IGNORECASE)
            if m:
                ten_kenh = m.group(1).strip()

        if ("http://" in s.lower() or "https://" in s.lower()) and not link_kenh:
            m = re.search(r"https?://\S+", s)
            if m:
                link_kenh = m.group(0).strip()

    if not link_kenh:
        for r in range(1, min(ws.max_row, 20) + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and ("http://" in v.lower() or "https://" in v.lower()):
                m = re.search(r"https?://\S+", v)
                if m:
                    link_kenh = m.group(0).strip()
                    break

    return {
        "contract_no": contract_no,
        "annex_no": annex_no,
        "ngay_ky_hop_dong": ngay_ky_hop_dong,
        "ngay_ky_phu_luc": ngay_ky_phu_luc,
        "ten_kenh": ten_kenh,
        "link_kenh": link_kenh,
    }


def _parse_import_table(ws) -> tuple[int, int]:
    header_row = None
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and isinstance(v2, str):
            if v1.strip().upper() == "STT" and "ID" in v2.upper():
                header_row = r
                break
    if not header_row:
        raise ValueError("Không tìm thấy dòng tiêu đề bảng tác phẩm (STT / ID Video)")
    return header_row, header_row + 1


def _norm_header_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())
    return s.strip().lower()


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm_header_cell(ws.cell(row=header_row, column=c).value)
        if not key:
            continue
        out.setdefault(key, c)
    return out


def _col(hmap: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = _norm_header_cell(n)
        if k in hmap:
            return hmap[k]
    return None


def _year_from_contract_no(contract_no: str) -> int:
    parts = (contract_no or "").split("/")
    if len(parts) >= 2 and parts[1].isdigit():
        return int(parts[1])
    return date.today().year
