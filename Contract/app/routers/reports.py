from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from app.auth import get_current_user, require_permission
from app.db import session_scope
from app.db_models import ContractRecordRow, UserRow, WorkRow
from app.utils import get_breadcrumbs


router = APIRouter()


_HEADER_FONT = Font(bold=True)


def _parse_iso_date(s: str | None, *, default: date) -> date:
    if not s:
        return default
    try:
        return date.fromisoformat(s)
    except Exception:
        return default


def _week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _group_key(d: date, group_by: str) -> str:
    g = (group_by or "day").strip().lower()
    if g == "day":
        return d.isoformat()
    if g == "week":
        ws = _week_start(d)
        we = ws + timedelta(days=6)
        return f"{ws.isoformat()}..{we.isoformat()}"
    if g == "month":
        return f"{d.year:04d}-{d.month:02d}"
    if g == "year":
        return f"{d.year:04d}"
    return d.isoformat()


def _date_from_work_imported_at(v: str | None) -> date | None:
    if not v:
        return None
    s = str(v).strip()
    if not s:
        return None

    # Common formats stored by works import
    # - YYYY-MM-DD HH:MM:SS
    # - YYYY-MM-DD
    try:
        if len(s) >= 10:
            return date.fromisoformat(s[:10])
    except Exception:
        pass

    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").date()
    except Exception:
        return None


def _xlsx_bytes(*, sheet_name: str, headers: list[str], rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT

    for r_idx, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c, value=v)

    bio = BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue()


def _build_report(
    *,
    user: UserRow,
    source: str | None,
    start: str | None,
    end: str | None,
    group_by: str | None,
    username: str | None,
) -> tuple[dict, list[dict], dict]:
    src = (source or "contracts").strip().lower()
    grp = (group_by or "month").strip().lower()

    today = date.today()
    default_start = date(today.year, 1, 1)
    start_d = _parse_iso_date(start, default=default_start)
    end_d = _parse_iso_date(end, default=today)

    if end_d < start_d:
        start_d, end_d = end_d, start_d

    is_admin_mod = user.role in ("admin", "mod")
    requested_user = (username or "").strip().lower()
    if is_admin_mod:
        user_filter = requested_user or None
    else:
        user_filter = user.username.strip().lower()

    buckets: dict[str, dict] = defaultdict(lambda: {"count": 0, "sum_value": 0})

    if src == "works":
        with session_scope() as db:
            q = db.query(WorkRow)
            if user_filter:
                q = q.filter(WorkRow.nguoi_thuc_hien == user_filter)
            db_rows = q.all()

        for r in db_rows:
            d = _date_from_work_imported_at(r.imported_at) or today
            if d < start_d or d > end_d:
                continue
            k = _group_key(d, grp)
            buckets[k]["count"] += 1

        out_rows = []
        for k in sorted(buckets.keys()):
            out_rows.append({"period": k, "works_count": buckets[k]["count"]})

        stats = {"total": sum(v["count"] for v in buckets.values())}

    else:
        with session_scope() as db:
            q = db.query(ContractRecordRow)
            if user_filter:
                q = q.filter(ContractRecordRow.nguoi_thuc_hien_email == user_filter)
            db_rows = q.all()

        for r in db_rows:
            if not r.ngay_lap_hop_dong:
                continue
            d = r.ngay_lap_hop_dong
            if d < start_d or d > end_d:
                continue
            k = _group_key(d, grp)
            buckets[k]["count"] += 1
            buckets[k]["sum_value"] += int(r.so_tien_value or 0)

        out_rows = []
        for k in sorted(buckets.keys()):
            out_rows.append(
                {
                    "period": k,
                    "contracts_count": buckets[k]["count"],
                    "total_value": buckets[k]["sum_value"],
                }
            )

        stats = {
            "total": sum(v["count"] for v in buckets.values()),
            "total_value": sum(v["sum_value"] for v in buckets.values()),
        }

    ctx = {
        "source": src,
        "group_by": grp,
        "start": start_d.isoformat(),
        "end": end_d.isoformat(),
        "user_filter": user_filter or "",
        "is_admin_mod": is_admin_mod,
    }

    return ctx, out_rows, stats


@router.get("/reports", response_class=HTMLResponse)
def reports_view(
    request: Request,
    source: str | None = None,
    start: str | None = None,
    end: str | None = None,
    group_by: str | None = None,
    username: str | None = None,
    user: UserRow = Depends(require_permission("reports.view")),
):
    templates = request.app.state.templates

    ctx, out_rows, stats = _build_report(
        user=user,
        source=source,
        start=start,
        end=end,
        group_by=group_by,
        username=username,
    )

    user_choices: list[str] = []
    if ctx["is_admin_mod"]:
        with session_scope() as db:
            user_choices = [u.username for u in db.query(UserRow).order_by(UserRow.username.asc()).all()]

    return templates.TemplateResponse(
        "reports.html",
        {
            "request": request,
            "title": "Báo cáo",
            "source": ctx["source"],
            "group_by": ctx["group_by"],
            "start": ctx["start"],
            "end": ctx["end"],
            "rows": out_rows,
            "stats": stats,
            "user_filter": ctx["user_filter"],
            "user_choices": user_choices,
            "is_admin_mod": ctx["is_admin_mod"],
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@router.get("/reports/export.xlsx")
def reports_export_excel(
    request: Request,
    source: str | None = None,
    start: str | None = None,
    end: str | None = None,
    group_by: str | None = None,
    username: str | None = None,
    user: UserRow = Depends(require_permission("reports.export")),
):
    ctx, out_rows, stats = _build_report(
        user=user,
        source=source,
        start=start,
        end=end,
        group_by=group_by,
        username=username,
    )

    if ctx["source"] == "works":
        headers = ["period", "works_count"]
        rows = [[r.get("period"), r.get("works_count")] for r in out_rows]
        sheet = "Works"
    else:
        headers = ["period", "contracts_count", "total_value"]
        rows = [[r.get("period"), r.get("contracts_count"), r.get("total_value")] for r in out_rows]
        sheet = "Contracts"

    rows.append([])
    if ctx["source"] == "works":
        rows.append(["TOTAL", int(stats.get("total") or 0)])
    else:
        rows.append(["TOTAL", int(stats.get("total") or 0), int(stats.get("total_value") or 0)])

    data = _xlsx_bytes(sheet_name=sheet, headers=headers, rows=rows)

    fname_user = (ctx.get("user_filter") or "all").replace("@", "_").replace(".", "_")
    filename = f"report_{ctx['source']}_{ctx['group_by']}_{ctx['start']}_{ctx['end']}_{fname_user}.xlsx"

    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
