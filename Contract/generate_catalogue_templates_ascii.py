from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _set(ws, cell: str, value: object, *, bold: bool = False, center: bool = False) -> None:
    c = ws[cell]
    c.value = value
    if bold:
        c.font = Font(bold=True)
    if center:
        c.alignment = Alignment(horizontal="center")


def _build_keys_sheet(wb: Workbook, *, keys: list[str]) -> None:
    ws = wb.create_sheet("KEYS")
    _set(ws, "A1", "placeholder", bold=True)
    _set(ws, "B1", "key", bold=True)

    for i, k in enumerate(keys, start=2):
        ws[f"A{i}"].value = f"<{k}>"
        ws[f"B{i}"].value = k

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40


def _common_layout(ws) -> None:
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 24


def build_contract_final(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Final"

    ws.merge_cells("A1:L1")
    _set(ws, "A1", "CATALOGUE TEMPLATE (CONTRACT)", bold=True, center=True)

    ws.merge_cells("A2:L2")
    _set(ws, "A2", "Contract No: <so_hop_dong_day_du>", center=True)

    ws.merge_cells("A3:L3")
    _set(ws, "A3", "Channel: <ten_kenh>", center=True)

    ws.merge_cells("A4:L4")
    _set(ws, "A4", "Channel link: <link_kenh>", center=True)

    _set(ws, "A6", "Company name:", bold=True)
    _set(ws, "B6", "<don_vi_ten>")
    _set(ws, "A7", "Address:", bold=True)
    _set(ws, "B7", "<don_vi_dia_chi>")
    _set(ws, "A8", "Phone:", bold=True)
    _set(ws, "B8", "<don_vi_dien_thoai>")
    _set(ws, "A9", "Representative:", bold=True)
    _set(ws, "B9", "<don_vi_nguoi_dai_dien>")
    _set(ws, "A10", "Email:", bold=True)
    _set(ws, "B10", "<don_vi_email>")

    headers = [
        "STT",
        "ID Video",
        "Code",
        "Work title",
        "Author",
        "Composer",
        "Lyricist",
        "Time range",
        "Duration",
        "Effective date",
        "Expiration date",
        "Usage type",
    ]
    start_row = 12
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = h
        cell.font = Font(bold=True)

    sample = [
        1,
        "<id_video>",
        "<code>",
        "<musical_work>",
        "<author>",
        "<composer>",
        "<lyricist>",
        "<time_range>",
        "<duration>",
        "<effective_date>",
        "<expiration_date>",
        "<usage_type>",
    ]
    for col, v in enumerate(sample, start=1):
        ws.cell(row=start_row + 1, column=col).value = v

    _set(ws, "J16", "Subtotal:", bold=True)
    _set(ws, "L16", "<so_tien_chua_GTGT>")
    _set(ws, "J17", "VAT <thue_percent>%:", bold=True)
    _set(ws, "L17", "<thue_GTGT>")
    _set(ws, "J18", "Total:", bold=True)
    _set(ws, "L18", "<so_tien>")
    _set(ws, "J19", "In words:", bold=True)
    _set(ws, "L19", "<so_tien_bang_chu>")

    _common_layout(ws)


def build_annex_final(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Final"

    ws.merge_cells("A1:L1")
    _set(ws, "A1", "CATALOGUE TEMPLATE (ANNEX)", bold=True, center=True)

    ws.merge_cells("A2:L2")
    _set(ws, "A2", "Contract: <so_hop_dong_day_du> | Annex: <so_phu_luc>", center=True)

    ws.merge_cells("A3:L3")
    _set(ws, "A3", "Contract date: <ngay_ky_hop_dong> | Annex date: <ngay_ky_phu_luc>", center=True)

    ws.merge_cells("A4:L4")
    _set(ws, "A4", "Channel: <ten_kenh> | Link: <link_kenh>", center=True)

    # reuse same body layout as contract
    build_contract_final(wb)
    ws = wb["Final"]
    ws["A1"].value = "CATALOGUE TEMPLATE (ANNEX)"
    ws["A2"].value = "Contract: <so_hop_dong_day_du> | Annex: <so_phu_luc>"
    ws["A3"].value = "Contract date: <ngay_ky_hop_dong> | Annex date: <ngay_ky_phu_luc>"
    ws["A4"].value = "Channel: <ten_kenh> | Link: <link_kenh>"


def write_template(*, out_path: Path, kind: str, force: bool) -> None:
    if out_path.exists() and not force:
        raise SystemExit(f"File exists: {out_path} (use --force)")

    wb = Workbook()

    if kind == "contract":
        build_contract_final(wb)
        keys = [
            "so_hop_dong_day_du",
            "contract_no",
            "don_vi_ten",
            "don_vi_dia_chi",
            "don_vi_dien_thoai",
            "don_vi_nguoi_dai_dien",
            "don_vi_chuc_vu",
            "don_vi_mst",
            "don_vi_email",
            "ten_kenh",
            "link_kenh",
            "so_tien_chua_GTGT",
            "thue_percent",
            "thue_GTGT",
            "so_tien",
            "so_tien_bang_chu",
            "id_video",
            "code",
            "musical_work",
            "author",
            "composer",
            "lyricist",
            "time_range",
            "duration",
            "effective_date",
            "expiration_date",
            "usage_type",
        ]
    elif kind == "annex":
        build_annex_final(wb)
        keys = [
            "so_hop_dong_day_du",
            "so_phu_luc",
            "ngay_ky_hop_dong",
            "ngay_ky_phu_luc",
            "don_vi_ten",
            "don_vi_dia_chi",
            "don_vi_dien_thoai",
            "don_vi_nguoi_dai_dien",
            "don_vi_chuc_vu",
            "don_vi_mst",
            "don_vi_email",
            "ten_kenh",
            "link_kenh",
            "so_tien_chua_GTGT",
            "thue_percent",
            "thue_GTGT",
            "so_tien",
            "so_tien_bang_chu",
            "id_video",
            "code",
            "musical_work",
            "author",
            "composer",
            "lyricist",
            "time_range",
            "duration",
            "effective_date",
            "expiration_date",
            "usage_type",
        ]
    else:
        raise SystemExit(f"Unknown kind: {kind}")

    _build_keys_sheet(wb, keys=keys)

    # remove default sheet if present
    if "Sheet" in wb.sheetnames:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--out-dir", default=str(Path(__file__).resolve().parent / "templates"))
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    write_template(
        out_path=out_dir / "TEMPLATE_DANH_MUC_HOP_DONG_PLACEHOLDERS.xlsx",
        kind="contract",
        force=args.force,
    )
    write_template(
        out_path=out_dir / "TEMPLATE_DANH_MUC_PHU_LUC_PLACEHOLDERS.xlsx",
        kind="annex",
        force=args.force,
    )
    print("OK: generated templates in", out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
