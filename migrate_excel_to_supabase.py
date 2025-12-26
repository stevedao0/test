"""
Migration script to import existing Excel data into Supabase database.
Run this once to migrate from Excel-based storage to Supabase.
"""
from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Add app to path
sys.path.insert(0, str(Path(__file__).parent))

from app.core.supabase import get_supabase
from app.services.auth import CurrentUser


def migrate_contracts_from_excel(excel_path: Path, year: int):
    """Migrate contracts and annexes from Excel to Supabase"""
    if not excel_path.exists():
        print(f"⏭️  Skipping {excel_path.name} (file not found)")
        return 0, 0

    print(f"📂 Reading {excel_path.name}...")
    wb = load_workbook(str(excel_path))
    ws = wb["Contracts"] if "Contracts" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"⏭️  Skipping {excel_path.name} (empty)")
        return 0, 0

    # Parse headers
    raw_headers = list(rows[0])
    headers = [h.strip() if isinstance(h, str) and h.strip() else None for h in raw_headers]

    supabase = get_supabase()
    contracts_count = 0
    annexes_count = 0

    # Create a fake user for migration (using service role, no RLS checks)
    # We'll set created_by to NULL since we don't have actual user IDs from Excel

    for row_data in rows[1:]:
        if not any(row_data):
            continue

        # Build row dict
        row_dict = {}
        for i in range(min(len(headers), len(row_data))):
            key = headers[i]
            if not key:
                continue
            row_dict[key] = row_data[i]

        contract_no = row_dict.get("contract_no")
        if not contract_no:
            continue

        annex_no = row_dict.get("annex_no")

        # Parse date
        ngay_lap = row_dict.get("ngay_lap_hop_dong")
        if isinstance(ngay_lap, datetime):
            ngay_lap = ngay_lap.date()
        elif isinstance(ngay_lap, str):
            try:
                ngay_lap = date.fromisoformat(ngay_lap)
            except:
                ngay_lap = date.today()
        elif not isinstance(ngay_lap, date):
            ngay_lap = date.today()

        # Prepare common fields
        data = {
            "contract_no": contract_no,
            "contract_year": row_dict.get("contract_year") or year,
            "ngay_lap_hop_dong": ngay_lap.isoformat(),
            "linh_vuc": row_dict.get("linh_vuc") or "Sao chép trực tuyến",
            "region_code": row_dict.get("region_code") or "HDQTGAN-PN",
            "field_code": row_dict.get("field_code") or "MR",
            "don_vi_ten": row_dict.get("don_vi_ten") or "",
            "don_vi_dia_chi": row_dict.get("don_vi_dia_chi"),
            "don_vi_dien_thoai": row_dict.get("don_vi_dien_thoai"),
            "don_vi_nguoi_dai_dien": row_dict.get("don_vi_nguoi_dai_dien"),
            "don_vi_chuc_vu": row_dict.get("don_vi_chuc_vu") or "Giám đốc",
            "don_vi_mst": row_dict.get("don_vi_mst"),
            "don_vi_email": row_dict.get("don_vi_email"),
            "so_cccd": row_dict.get("so_CCCD") or row_dict.get("so_cccd"),
            "ngay_cap_cccd": row_dict.get("ngay_cap_CCCD") or row_dict.get("ngay_cap_cccd"),
            "kenh_ten": row_dict.get("kenh_ten"),
            "kenh_id": row_dict.get("kenh_id"),
            "nguoi_thuc_hien_email": row_dict.get("nguoi_thuc_hien_email"),
            "so_tien_chua_gtgt_value": _safe_int(row_dict.get("so_tien_chua_GTGT_value")),
            "so_tien_chua_gtgt_text": row_dict.get("so_tien_chua_GTGT_text"),
            "thue_percent": _safe_float(row_dict.get("thue_percent")),
            "thue_gtgt_value": _safe_int(row_dict.get("thue_GTGT_value")),
            "thue_gtgt_text": row_dict.get("thue_GTGT_text"),
            "so_tien_value": _safe_int(row_dict.get("so_tien_value")),
            "so_tien_text": row_dict.get("so_tien_text"),
            "so_tien_bang_chu": row_dict.get("so_tien_bang_chu"),
            "docx_path": row_dict.get("docx_path"),
            "catalogue_path": row_dict.get("catalogue_path"),
        }

        try:
            if not annex_no:
                # This is a contract
                # Check if already exists
                existing = supabase.table("contracts").select("id").eq("contract_no", contract_no).maybeSingle().execute()
                if existing.data:
                    print(f"  ⏭️  Contract {contract_no} already exists")
                    continue

                supabase.table("contracts").insert(data).execute()
                contracts_count += 1
                print(f"  ✅ Migrated contract: {contract_no}")
            else:
                # This is an annex
                # Find parent contract
                parent = supabase.table("contracts").select("id").eq("contract_no", contract_no).maybeSingle().execute()
                if not parent.data:
                    print(f"  ⚠️  Parent contract {contract_no} not found for annex {annex_no}, skipping")
                    continue

                # Check if annex already exists
                existing = supabase.table("annexes").select("id").eq("contract_id", parent.data["id"]).eq("annex_no", annex_no).maybeSingle().execute()
                if existing.data:
                    print(f"  ⏭️  Annex {contract_no}/{annex_no} already exists")
                    continue

                annex_data = data.copy()
                del annex_data["contract_no"]
                del annex_data["contract_year"]
                del annex_data["region_code"]
                del annex_data["field_code"]
                annex_data["contract_id"] = parent.data["id"]
                annex_data["annex_no"] = annex_no
                annex_data["ngay_ky_phu_luc"] = ngay_lap.isoformat()

                supabase.table("annexes").insert(annex_data).execute()
                annexes_count += 1
                print(f"  ✅ Migrated annex: {contract_no}/{annex_no}")
        except Exception as e:
            print(f"  ❌ Error migrating {contract_no}/{annex_no or 'main'}: {e}")
            continue

    return contracts_count, annexes_count


def _safe_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    if isinstance(val, str):
        try:
            return int(val.replace(",", "").replace(".", ""))
        except:
            return None
    return None


def _safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.replace(",", "."))
        except:
            return None
    return None


def migrate_works_from_excel(excel_path: Path, year: int):
    """Migrate works from Excel to Supabase"""
    if not excel_path.exists():
        print(f"⏭️  Skipping {excel_path.name} (file not found)")
        return 0

    print(f"📂 Reading {excel_path.name}...")
    wb = load_workbook(str(excel_path))
    ws = wb["Works"] if "Works" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"⏭️  Skipping {excel_path.name} (empty)")
        return 0

    # Parse headers
    raw_headers = list(rows[0])
    headers = [h.strip() if isinstance(h, str) and h.strip() else None for h in raw_headers]

    supabase = get_supabase()
    works_count = 0

    for row_data in rows[1:]:
        if not any(row_data):
            continue

        # Build row dict
        row_dict = {}
        for i in range(min(len(headers), len(row_data))):
            key = headers[i]
            if not key:
                continue
            row_dict[key] = row_data[i]

        contract_no = row_dict.get("contract_no")
        annex_no = row_dict.get("annex_no")

        if not contract_no:
            continue

        # Find contract
        contract = supabase.table("contracts").select("id").eq("contract_no", contract_no).maybeSingle().execute()
        if not contract.data:
            print(f"  ⚠️  Contract {contract_no} not found, skipping work")
            continue

        contract_id = contract.data["id"]
        annex_id = None

        if annex_no:
            annex = supabase.table("annexes").select("id").eq("contract_id", contract_id).eq("annex_no", annex_no).maybeSingle().execute()
            if annex.data:
                annex_id = annex.data["id"]

        work_data = {
            "contract_id": contract_id,
            "annex_id": annex_id,
            "stt": _safe_int(row_dict.get("stt")) or 0,
            "id_link": row_dict.get("id_link"),
            "youtube_url": row_dict.get("youtube_url"),
            "id_work": row_dict.get("id_work"),
            "musical_work": row_dict.get("musical_work"),
            "author": row_dict.get("author"),
            "composer": row_dict.get("composer"),
            "lyricist": row_dict.get("lyricist"),
            "time_range": row_dict.get("time_range"),
            "duration": row_dict.get("duration"),
            "effective_date": row_dict.get("effective_date"),
            "expiration_date": row_dict.get("expiration_date"),
            "usage_type": row_dict.get("usage_type"),
            "royalty_rate": str(row_dict.get("royalty_rate")) if row_dict.get("royalty_rate") else None,
            "note": row_dict.get("note"),
            "imported_at": row_dict.get("imported_at"),
            "nguoi_thuc_hien": row_dict.get("nguoi_thuc_hien"),
        }

        try:
            supabase.table("works").insert(work_data).execute()
            works_count += 1
            if works_count % 100 == 0:
                print(f"  📝 Migrated {works_count} works...")
        except Exception as e:
            print(f"  ❌ Error migrating work STT {row_dict.get('stt')}: {e}")
            continue

    print(f"  ✅ Migrated {works_count} works")
    return works_count


def main():
    print("🚀 Starting Excel to Supabase migration...\n")

    storage_dir = Path("storage/excel")

    total_contracts = 0
    total_annexes = 0
    total_works = 0

    # Migrate contracts and annexes from each year
    for year in [2025]:  # Add more years as needed
        print(f"\n📅 Migrating year {year}...")
        contracts_path = storage_dir / f"contracts_{year}.xlsx"
        c, a = migrate_contracts_from_excel(contracts_path, year)
        total_contracts += c
        total_annexes += a

        # Migrate works
        works_path = storage_dir / f"works_contract_{year}.xlsx"
        w = migrate_works_from_excel(works_path, year)
        total_works += w

    print(f"\n✨ Migration complete!")
    print(f"   📄 Contracts: {total_contracts}")
    print(f"   📎 Annexes: {total_annexes}")
    print(f"   🎵 Works: {total_works}")
    print(f"\n💡 You can now remove the Excel files and excel_store.py")


if __name__ == "__main__":
    main()
