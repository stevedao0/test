from __future__ import annotations

from io import BytesIO
from datetime import date, datetime

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

from app.config import STORAGE_EXCEL_DIR
from app.services.excel_store import append_works_rows
from app.utils.formatters import (
    format_ddmmyyyy,
    extract_channel_id,
    extract_video_id,
    normalize_time_range,
    normalize_hhmmss,
)

router = APIRouter()


def get_breadcrumbs(path: str):
    breadcrumbs = [{"label": "Trang chủ", "url": "/"}]

    if "/contracts" in path:
        breadcrumbs.append({"label": "Hợp đồng", "url": "/contracts"})
        if "/new" in path:
            breadcrumbs.append({"label": "Tạo mới", "url": None})
    elif "/annexes" in path:
        breadcrumbs.append({"label": "Phụ lục", "url": "/annexes"})
        if "/new" in path:
            breadcrumbs.append({"label": "Tạo mới", "url": None})

    return breadcrumbs


def parse_import_metadata(ws) -> dict:
    import re

    contract_no = ""
    annex_no = ""
    ngay_ky_hop_dong = ""
    ngay_ky_phu_luc = ""
    ten_kenh = ""
    link_kenh = ""

    for r in range(1, min(ws.max_row, 15) + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        s = " ".join(v.split())

        if "HỢP ĐỒNG SỐ" in s.upper() and not contract_no:
            m = re.search(r"HỢP\s*ĐỒNG\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                contract_no = m.group(1).strip()
            m2 = re.search(r"NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_hop_dong = format_ddmmyyyy(m2.group(1))

        if "PHỤ LỤC SỐ" in s.upper() and not annex_no:
            m = re.search(r"PHỤ\s*LỤC\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                annex_no = m.group(1).strip()
            m2 = re.search(r"PHỤ\s*LỤC\s*SỐ\s*[^\s]+\s*NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_phu_luc = format_ddmmyyyy(m2.group(1))

        if "YOUTUBE" in s.upper() and not ten_kenh:
            m = re.search(r"YOUTUBE\s+(.+)$", s, re.IGNORECASE)
            if m:
                ten_kenh = m.group(1).strip()

        if ("http://" in s.lower() or "https://" in s.lower()) and not link_kenh:
            m = re.search(r"https?://\S+", s)
            if m:
                link_kenh = m.group(0).strip()

    if not link_kenh:
        for r in range(1, min(ws.max_row, 20) + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and ("http://" in v.lower() or "https://" in v.lower()):
                m = re.search(r"https?://\S+", v)
                if m:
                    link_kenh = m.group(0).strip()
                    break

    return {
        "contract_no": contract_no,
        "annex_no": annex_no,
        "ngay_ky_hop_dong": ngay_ky_hop_dong,
        "ngay_ky_phu_luc": ngay_ky_phu_luc,
        "ten_kenh": ten_kenh,
        "link_kenh": link_kenh,
    }


def parse_import_table(ws) -> tuple[int, int]:
    header_row = None
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and isinstance(v2, str):
            if v1.strip().upper() == "STT" and "ID" in v2.upper():
                header_row = r
                break
    if not header_row:
        raise ValueError("Không tìm thấy dòng tiêu đề bảng tác phẩm (STT / ID Video)")
    return header_row, header_row + 1


def norm_header_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())
    return s.strip().lower()


def build_header_map(ws, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = norm_header_cell(ws.cell(row=header_row, column=c).value)
        if not key:
            continue
        out.setdefault(key, c)
    return out


def col(hmap: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = norm_header_cell(n)
        if k in hmap:
            return hmap[k]
    return None


def year_from_contract_no(contract_no: str) -> int:
    parts = (contract_no or "").split("/")
    if len(parts) >= 2 and parts[1].isdigit():
        return int(parts[1])
    return date.today().year


@router.get("/works/import", response_class=HTMLResponse)
def works_import_form(request: Request, error: str | None = None, message: str | None = None):
    from pathlib import Path
    templates_dir = Path("app/web_templates")
    templates = Jinja2Templates(directory=str(templates_dir))

    return templates.TemplateResponse(
        "works_import.html",
        {
            "request": request,
            "title": "Import danh sách tác phẩm",
            "error": error,
            "message": message,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@router.post("/works/import")
async def works_import_submit(
    request: Request,
    import_file: UploadFile = File(...),
    nguoi_thuc_hien: str = Form(""),
):
    try:
        data = await import_file.read()
        wb = load_workbook(filename=BytesIO(data), data_only=False)
        ws = wb[wb.sheetnames[0]]

        meta = parse_import_metadata(ws)
        contract_no = meta.get("contract_no", "")
        annex_no = meta.get("annex_no", "")
        year = year_from_contract_no(contract_no)

        uploaded_filename = import_file.filename or ""
        catalogue_dir = STORAGE_EXCEL_DIR / str(year)
        existing_catalogue_path = None
        if uploaded_filename and catalogue_dir.exists():
            potential_path = catalogue_dir / uploaded_filename
            if potential_path.exists():
                existing_catalogue_path = potential_path

        id_channel = extract_channel_id(meta.get("link_kenh", ""))

        header_row, start_row = parse_import_table(ws)
        hmap = build_header_map(ws, header_row)

        c_stt = col(hmap, "stt") or 1
        c_id_video = col(hmap, "id video") or 2
        c_code = col(hmap, "code") or 3
        c_title = col(hmap, "tên tác phẩm") or 4
        c_author = col(hmap, "tên tác giả") or 5
        c_composer = col(hmap, "tên tác giả nhạc") or 6
        c_lyricist = col(hmap, "tên tác giả lời") or 7
        c_time_range = col(hmap, "thời gian") or 8
        c_duration = col(hmap, "thời lượng") or 9
        c_effective = col(hmap, "ngày bắt đầu") or 10
        c_expiration = col(hmap, "thời hạn kết thúc") or 11
        c_usage = col(hmap, "hình thức sử dụng") or 12
        c_rate = col(hmap, "mức nhuận bút chưa thuế gtgt (vnđ)") or 13
        c_note = col(hmap, "ghi chú")

        imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        out_rows: list[dict] = []

        for r in range(start_row, ws.max_row + 1):
            stt = ws.cell(row=r, column=c_stt).value
            if stt is None:
                continue

            if isinstance(stt, str) and stt.strip().lower().startswith("cộng"):
                break

            try:
                stt_int = int(str(stt).strip())
            except Exception:
                continue

            id_video = ws.cell(row=r, column=c_id_video).value
            id_link = extract_video_id(id_video)
            youtube_url = f"https://www.youtube.com/watch?v={id_link}" if id_link else ""

            id_work = ws.cell(row=r, column=c_code).value
            musical_work = ws.cell(row=r, column=c_title).value
            author = ws.cell(row=r, column=c_author).value
            composer = ws.cell(row=r, column=c_composer).value
            lyricist = ws.cell(row=r, column=c_lyricist).value

            time_range_raw = ws.cell(row=r, column=c_time_range).value
            duration_raw = ws.cell(row=r, column=c_duration).value
            effective_date_raw = ws.cell(row=r, column=c_effective).value
            expiration_date_raw = ws.cell(row=r, column=c_expiration).value
            usage_type = ws.cell(row=r, column=c_usage).value
            royalty_rate = ws.cell(row=r, column=c_rate).value
            note = ws.cell(row=r, column=c_note).value if c_note else ""

            time_range = ""
            duration = ""
            if time_range_raw is not None and str(time_range_raw).strip():
                time_range = normalize_time_range(str(time_range_raw))
            if duration_raw is not None and str(duration_raw).strip():
                duration = normalize_hhmmss(str(duration_raw))

            effective_date = format_ddmmyyyy(effective_date_raw)
            expiration_date = format_ddmmyyyy(expiration_date_raw)

            rr = {
                "year": year,
                "contract_no": contract_no,
                "annex_no": annex_no or "",
                "ngay_ky_hop_dong": meta.get("ngay_ky_hop_dong", ""),
                "ngay_ky_phu_luc": meta.get("ngay_ky_phu_luc", ""),
                "nguoi_thuc_hien": (nguoi_thuc_hien or "").strip(),
                "ten_kenh": meta.get("ten_kenh", ""),
                "id_channel": id_channel,
                "link_kenh": meta.get("link_kenh", ""),
                "stt": stt_int,
                "id_link": id_link,
                "youtube_url": youtube_url,
                "id_work": id_work,
                "musical_work": musical_work,
                "author": author,
                "composer": composer,
                "lyricist": lyricist,
                "time_range": time_range,
                "duration": duration,
                "effective_date": effective_date,
                "expiration_date": expiration_date,
                "usage_type": usage_type,
                "royalty_rate": royalty_rate,
                "note": note,
                "imported_at": imported_at,
            }
            out_rows.append(rr)

        if not contract_no:
            raise ValueError("Không đọc được số hợp đồng trong file import")

        out_path = STORAGE_EXCEL_DIR / f"works_contract_{year}.xlsx"
        append_works_rows(excel_path=out_path, rows=out_rows)

        if existing_catalogue_path:
            existing_catalogue_path.write_bytes(data)
            return RedirectResponse(
                url=f"/works/import?message=Đã import {len(out_rows)} dòng vào {out_path.name} và cập nhật file danh mục {uploaded_filename}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/works/import?message=Đã import {len(out_rows)} dòng vào {out_path.name}",
            status_code=303,
        )
    except Exception as e:
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(url=f"/works/import?error={msg}", status_code=303)
