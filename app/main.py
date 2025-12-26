from __future__ import annotations

import re
import traceback
from io import BytesIO
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

from app.config import (
    ANNEX_TEMPLATE_PATH,
    ANNEX_CATALOGUE_TEMPLATE_PATH,
    CATALOGUE_TEMPLATE_PATH,
    DOCX_TEMPLATE_PATH,
    STORAGE_DOCX_DIR,
    STORAGE_EXCEL_DIR,
    UI_STATIC_DIR,
    UI_TEMPLATES_DIR,
    WEB_TEMPLATES_DIR,
)
from app.documents.naming import build_docx_filename
from app.models import ContractCreate, ContractRecord
# from app.services.annex_store import append_annex_row  # No longer needed - annexes saved to contracts Excel
from app.services.docx_renderer import date_parts, render_contract_docx
from app.services.excel_store import (
    append_contract_row,
    append_works_rows,
    delete_contract_row,
    export_catalogue_excel,
    read_contracts,
    update_contract_row,
)


app = FastAPI()


def _pick_existing_dir(primary: Path, fallback: Path) -> Path:
    try:
        if primary.exists():
            # Consider it usable if it has any entries
            for _ in primary.iterdir():
                return primary
    except Exception:
        pass
    return fallback


def _pick_templates_dir(primary: Path, fallback: Path) -> Path:
    # Only pick the new UI templates dir once the essential templates are present.
    try:
        if primary.exists() and (primary / "document_form.html").exists():
            return primary
    except Exception:
        pass
    return fallback


def _pick_static_dir(primary: Path, fallback: Path) -> Path:
    # Only pick the new UI static dir once core assets are present.
    try:
        if primary.exists() and (primary / "css" / "main.css").exists():
            return primary
    except Exception:
        pass
    return fallback


_templates_dir = _pick_templates_dir(UI_TEMPLATES_DIR, WEB_TEMPLATES_DIR)
_static_dir = _pick_static_dir(UI_STATIC_DIR, Path("app/static"))

app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")
templates = Jinja2Templates(directory=str(_templates_dir))

REGION_CODE = "HĐQTGAN-PN"
FIELD_CODE = "MR"
FIELD_NAME = "Sao chép trực tuyến"


def _format_ddmmyyyy(v) -> str:
    if v is None:
        return ""

    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    if not s:
        return ""

    normalized = s.replace("-", "/").replace(".", "/")

    # Accept d/m/yyyy or dd/mm/yyyy (also with '-')
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", normalized)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{d:02d}/{mo:02d}/{y:04d}"

    # If it's not parseable, keep original string (do not hard-fail on import)
    return s


def _extract_channel_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    m = re.search(r"(UC[0-9A-Za-z_-]{10,})", s)
    if m:
        return m.group(1)
    return ""


def _extract_video_id(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""

    if s.startswith("="):
        m = re.search(r"watch\?v=([0-9A-Za-z_-]{6,})", s)
        if m:
            return m.group(1)
        m = re.search(r"HYPERLINK\(\"[^\"]*youtu(?:\.be/|be\.com/watch\?v=)([^\"&?#]+)", s, re.IGNORECASE)
        if m:
            return m.group(1)

    m = re.search(r"watch\?v=([0-9A-Za-z_-]{6,})", s)
    if m:
        return m.group(1)
    m = re.search(r"youtu\.be/([0-9A-Za-z_-]{6,})", s)
    if m:
        return m.group(1)

    if re.fullmatch(r"[0-9A-Za-z_-]{6,}", s):
        return s
    return ""


def _normalize_hhmmss(s: str) -> str:
    """Normalize time into hh:mm:ss.

    Accepts:
    - hh:mm:ss
    - mm:ss (will become 00:mm:ss)
    """

    t = (s or "").strip()
    if not t:
        return ""

    parts = [p.strip() for p in t.split(":")]
    if len(parts) == 2:
        hh = 0
        mm = int(parts[0])
        ss = int(parts[1])
    elif len(parts) == 3:
        hh = int(parts[0])
        mm = int(parts[1])
        ss = int(parts[2])
    else:
        raise ValueError("Thời lượng/thời gian phải theo định dạng hh:mm:ss hoặc mm:ss")

    if mm < 0 or mm >= 60 or ss < 0 or ss >= 60 or hh < 0:
        raise ValueError("Thời lượng/thời gian không hợp lệ")

    return f"{hh:02d}:{mm:02d}:{ss:02d}"


def _normalize_time_range(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""

    # accept '-' or '–'
    tt = t.replace("–", "-")
    parts = [p.strip() for p in tt.split("-")]
    if len(parts) != 2:
        raise ValueError("Thời gian phải theo định dạng hh:mm:ss - hh:mm:ss (hoặc mm:ss - mm:ss)")

    start = _normalize_hhmmss(parts[0])
    end = _normalize_hhmmss(parts[1])
    return f"{start} - {end}"


def _parse_import_metadata(ws) -> dict:
    contract_no = ""
    annex_no = ""
    ngay_ky_hop_dong = ""
    ngay_ky_phu_luc = ""
    ten_kenh = ""
    link_kenh = ""

    for r in range(1, min(ws.max_row, 15) + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        s = " ".join(v.split())

        if "HỢP ĐỒNG SỐ" in s.upper() and not contract_no:
            m = re.search(r"HỢP\s*ĐỒNG\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                contract_no = m.group(1).strip()
            m2 = re.search(r"NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_hop_dong = _format_ddmmyyyy(m2.group(1))

        if "PHỤ LỤC SỐ" in s.upper() and not annex_no:
            m = re.search(r"PHỤ\s*LỤC\s*SỐ\s*([^\s]+)", s, re.IGNORECASE)
            if m:
                annex_no = m.group(1).strip()
            m2 = re.search(r"PHỤ\s*LỤC\s*SỐ\s*[^\s]+\s*NGÀY\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.IGNORECASE)
            if m2:
                ngay_ky_phu_luc = _format_ddmmyyyy(m2.group(1))

        if "YOUTUBE" in s.upper() and not ten_kenh:
            m = re.search(r"YOUTUBE\s+(.+)$", s, re.IGNORECASE)
            if m:
                ten_kenh = m.group(1).strip()

        if ("http://" in s.lower() or "https://" in s.lower()) and not link_kenh:
            m = re.search(r"https?://\S+", s)
            if m:
                link_kenh = m.group(0).strip()

    if not link_kenh:
        for r in range(1, min(ws.max_row, 20) + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and ("http://" in v.lower() or "https://" in v.lower()):
                m = re.search(r"https?://\S+", v)
                if m:
                    link_kenh = m.group(0).strip()
                    break

    return {
        "contract_no": contract_no,
        "annex_no": annex_no,
        "ngay_ky_hop_dong": ngay_ky_hop_dong,
        "ngay_ky_phu_luc": ngay_ky_phu_luc,
        "ten_kenh": ten_kenh,
        "link_kenh": link_kenh,
    }


def _parse_import_table(ws) -> tuple[int, int]:
    header_row = None
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and isinstance(v2, str):
            if v1.strip().upper() == "STT" and "ID" in v2.upper():
                header_row = r
                break
    if not header_row:
        raise ValueError("Không tìm thấy dòng tiêu đề bảng tác phẩm (STT / ID Video)")
    return header_row, header_row + 1


def _norm_header_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())
    return s.strip().lower()


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    # map normalized header -> column index (1-based)
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = _norm_header_cell(ws.cell(row=header_row, column=c).value)
        if not key:
            continue
        # keep first occurrence
        out.setdefault(key, c)
    return out


def _col(hmap: dict[str, int], *names: str) -> int | None:
    for n in names:
        k = _norm_header_cell(n)
        if k in hmap:
            return hmap[k]
    return None


def _year_from_contract_no(contract_no: str) -> int:
    parts = (contract_no or "").split("/")
    if len(parts) >= 2 and parts[1].isdigit():
        return int(parts[1])
    return date.today().year


@app.get("/works/import", response_class=HTMLResponse)
def works_import_form(request: Request, error: str | None = None, message: str | None = None):
    return templates.TemplateResponse(
        "works_import.html",
        {
            "request": request,
            "title": "Import danh sách tác phẩm",
            "error": error,
            "message": message,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/works/import")
async def works_import_submit(
    request: Request,
    import_file: UploadFile = File(...),
    nguoi_thuc_hien: str = Form(""),
):
    try:
        data = await import_file.read()
        wb = load_workbook(filename=BytesIO(data), data_only=False)
        ws = wb[wb.sheetnames[0]]

        meta = _parse_import_metadata(ws)
        contract_no = meta.get("contract_no", "")
        annex_no = meta.get("annex_no", "")
        year = _year_from_contract_no(contract_no)

        id_channel = _extract_channel_id(meta.get("link_kenh", ""))

        header_row, start_row = _parse_import_table(ws)
        hmap = _build_header_map(ws, header_row)

        c_stt = _col(hmap, "stt") or 1
        c_id_video = _col(hmap, "id video") or 2
        c_code = _col(hmap, "code") or 3
        c_title = _col(hmap, "tên tác phẩm") or 4
        c_author = _col(hmap, "tên tác giả") or 5
        c_composer = _col(hmap, "tên tác giả nhạc") or 6
        c_lyricist = _col(hmap, "tên tác giả lời") or 7
        c_time_range = _col(hmap, "thời gian") or 8
        c_duration = _col(hmap, "thời lượng") or 9
        c_effective = _col(hmap, "ngày bắt đầu") or 10
        c_expiration = _col(hmap, "thời hạn kết thúc") or 11
        c_usage = _col(hmap, "hình thức sử dụng") or 12
        c_rate = _col(hmap, "mức nhuận bút chưa thuế gtgt (vnđ)") or 13
        c_note = _col(hmap, "ghi chú")

        imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        out_rows: list[dict] = []

        for r in range(start_row, ws.max_row + 1):
            stt = ws.cell(row=r, column=c_stt).value
            if stt is None:
                continue

            if isinstance(stt, str) and stt.strip().lower().startswith("cộng"):
                break

            try:
                stt_int = int(str(stt).strip())
            except Exception:
                continue

            id_video = ws.cell(row=r, column=c_id_video).value
            id_link = _extract_video_id(id_video)
            youtube_url = f"https://www.youtube.com/watch?v={id_link}" if id_link else ""

            id_work = ws.cell(row=r, column=c_code).value
            musical_work = ws.cell(row=r, column=c_title).value
            author = ws.cell(row=r, column=c_author).value
            composer = ws.cell(row=r, column=c_composer).value
            lyricist = ws.cell(row=r, column=c_lyricist).value

            time_range_raw = ws.cell(row=r, column=c_time_range).value
            duration_raw = ws.cell(row=r, column=c_duration).value
            effective_date_raw = ws.cell(row=r, column=c_effective).value
            expiration_date_raw = ws.cell(row=r, column=c_expiration).value
            usage_type = ws.cell(row=r, column=c_usage).value
            royalty_rate = ws.cell(row=r, column=c_rate).value
            note = ws.cell(row=r, column=c_note).value if c_note else ""

            time_range = ""
            duration = ""
            if time_range_raw is not None and str(time_range_raw).strip():
                time_range = _normalize_time_range(str(time_range_raw))
            if duration_raw is not None and str(duration_raw).strip():
                duration = _normalize_hhmmss(str(duration_raw))

            effective_date = _format_ddmmyyyy(effective_date_raw)
            expiration_date = _format_ddmmyyyy(expiration_date_raw)

            rr = {
                "year": year,
                "contract_no": contract_no,
                "annex_no": annex_no or "",
                "ngay_ky_hop_dong": meta.get("ngay_ky_hop_dong", ""),
                "ngay_ky_phu_luc": meta.get("ngay_ky_phu_luc", ""),
                "nguoi_thuc_hien": (nguoi_thuc_hien or "").strip(),
                "ten_kenh": meta.get("ten_kenh", ""),
                "id_channel": id_channel,
                "link_kenh": meta.get("link_kenh", ""),
                "stt": stt_int,
                "id_link": id_link,
                "youtube_url": youtube_url,
                "id_work": id_work,
                "musical_work": musical_work,
                "author": author,
                "composer": composer,
                "lyricist": lyricist,
                "time_range": time_range,
                "duration": duration,
                "effective_date": effective_date,
                "expiration_date": expiration_date,
                "usage_type": usage_type,
                "royalty_rate": royalty_rate,
                "note": note,
                "imported_at": imported_at,
            }
            out_rows.append(rr)

        if not contract_no:
            raise ValueError("Không đọc được số hợp đồng trong file import")

        out_path = STORAGE_EXCEL_DIR / f"works_contract_{year}.xlsx"
        append_works_rows(excel_path=out_path, rows=out_rows)

        return RedirectResponse(
            url=f"/works/import?message=Đã import {len(out_rows)} dòng vào {out_path.name}",
            status_code=303,
        )
    except Exception as e:
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(url=f"/works/import?error={msg}", status_code=303)


def _clean_opt(s: Optional[str]) -> str:
    if s is None:
        return ""
    return s.strip()


def _split_multi_values(s: str) -> list[str]:
    if not s:
        return []
    normalized = s.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace(",", ";").replace("\n", ";")
    parts = [p.strip() for p in normalized.split(";")]
    return [p for p in parts if p]


def normalize_multi_emails(s: str) -> str:
    parts = _split_multi_values(_clean_opt(s))
    if not parts:
        return ""
    return "; ".join(parts)


def normalize_multi_phones(s: str) -> str:
    parts = _split_multi_values(_clean_opt(s))
    if not parts:
        return ""

    out: list[str] = []
    for p in parts:
        compact = "".join([ch for ch in p if ch.isdigit()])
        if compact and len(compact) in (10, 11) and compact.startswith("0"):
            out.append(compact)
        else:
            out.append(" ".join(p.split()))
    return "; ".join(out)


def normalize_youtube_channel_input(raw: str) -> tuple[str, str]:
    """Accept UC... or any youtube channel URL and return (channel_id, full_url)."""

    s = (raw or "").strip()
    if not s:
        return "", ""

    # If user pastes a full URL, extract UC id
    m = re.search(r"(UC[0-9A-Za-z_-]{10,})", s)
    channel_id = m.group(1) if m else s
    if not channel_id.startswith("UC"):
        # Not a UC id, keep as-is but still avoid double prefixing
        return channel_id, s

    return channel_id, f"https://www.youtube.com/channel/{channel_id}"


def normalize_money_to_int(s: str) -> int:
    # Accept: 15,600,000 or 15600000 or '15,600,000 VNĐ'
    raw = s.strip()
    raw = raw.replace("VNĐ", "").replace("VND", "").strip()
    digits = re.sub(r"[^0-9]", "", raw)
    if not digits:
        raise ValueError("Số tiền không hợp lệ")
    return int(digits)


def format_money_vnd(n: int) -> str:
    return f"{n:,} VNĐ".replace(",", ",")


def format_money_number(n: int) -> str:
    return f"{n:,}".replace(",", ",")


def _vi_three_digits(n: int, *, full: bool) -> str:
    ones = [
        "không",
        "một",
        "hai",
        "ba",
        "bốn",
        "năm",
        "sáu",
        "bảy",
        "tám",
        "chín",
    ]

    tram = n // 100
    chuc = (n % 100) // 10
    donvi = n % 10

    parts: list[str] = []

    if tram > 0 or full:
        parts.append(f"{ones[tram]} trăm")

    if chuc == 0:
        if donvi != 0 and (tram > 0 or full):
            parts.append("lẻ")
    elif chuc == 1:
        parts.append("mười")
    else:
        parts.append(f"{ones[chuc]} mươi")

    if donvi == 0:
        return " ".join(parts).strip()
    if donvi == 1:
        if chuc >= 2:
            parts.append("mốt")
        else:
            parts.append("một")
        return " ".join(parts).strip()
    if donvi == 4:
        if chuc >= 2:
            parts.append("tư")
        else:
            parts.append("bốn")
        return " ".join(parts).strip()
    if donvi == 5:
        if chuc >= 1:
            parts.append("lăm")
        else:
            parts.append("năm")
        return " ".join(parts).strip()

    parts.append(ones[donvi])
    return " ".join(parts).strip()


def money_to_vietnamese_words(n: int) -> str:
    if n == 0:
        return "Không đồng"
    if n < 0:
        return "Âm " + money_to_vietnamese_words(-n)

    units = [
        (10**9, "tỷ"),
        (10**6, "triệu"),
        (10**3, "nghìn"),
    ]

    parts: list[str] = []
    remainder = n
    started = False

    for base, name in units:
        block = remainder // base
        remainder = remainder % base
        if block == 0:
            continue
        started = True
        parts.append(f"{_vi_three_digits(block, full=False)} {name}")

    if remainder > 0:
        parts.append(_vi_three_digits(remainder, full=started))

    text = " ".join(p for p in parts if p).strip()
    text = text[0].upper() + text[1:] if text else text
    return f"{text} đồng"


def get_breadcrumbs(path: str):
    breadcrumbs = [{"label": "Trang chủ", "url": "/"}]

    if "/contracts" in path:
        breadcrumbs.append({"label": "Hợp đồng", "url": "/contracts"})
        if "/new" in path:
            breadcrumbs.append({"label": "Tạo mới", "url": None})
    elif "/annexes" in path:
        breadcrumbs.append({"label": "Phụ lục", "url": "/annexes"})
        if "/new" in path:
            breadcrumbs.append({"label": "Tạo mới", "url": None})

    return breadcrumbs


@app.get("/", response_class=HTMLResponse)
def home() -> RedirectResponse:
    return RedirectResponse(url="/documents/new")


@app.get("/contracts/new", response_class=HTMLResponse)
def contract_form(request: Request, error: str | None = None):
    url = f"/documents/new?doc_type=contract"
    if error:
        url += f"&error={error}"
    return RedirectResponse(url=url)


@app.post("/contracts")
def create_contract(
    request: Request,
    ngay_lap_hop_dong: str = Form(...),
    so_hop_dong_4: str = Form(...),
    linh_vuc: str = Form(FIELD_NAME),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form("Giám đốc"),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    so_CCCD: str = Form(""),
    ngay_cap_CCCD: str = Form(""),
    nguoi_thuc_hien_email: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form(""),
):
    try:
        channel_id, channel_link = normalize_youtube_channel_input(kenh_id)

        linh_vuc_value = _clean_opt(linh_vuc) or FIELD_NAME

        # Money option B
        pre_vat_value: Optional[int] = None
        pre_vat_text = ""
        pre_vat_number = ""
        vat_percent_value: Optional[float] = None
        vat_value: Optional[int] = None
        vat_text = ""
        vat_number = ""
        total_value: Optional[int] = None
        total_text = ""
        total_number = ""
        total_words = ""

        if _clean_opt(so_tien_chua_GTGT):
            pre_vat_value = normalize_money_to_int(_clean_opt(so_tien_chua_GTGT))
            pre_vat_text = format_money_vnd(pre_vat_value)
            pre_vat_number = format_money_number(pre_vat_value)

            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_value = float(pct_raw.replace(",", "."))
            if vat_percent_value < 0:
                raise ValueError("Thuế GTGT không hợp lệ")

            vat_value = int(round(pre_vat_value * vat_percent_value / 100.0))
            vat_text = format_money_vnd(vat_value)
            vat_number = format_money_number(vat_value)

            total_value = pre_vat_value + vat_value
            total_text = format_money_vnd(total_value)
            total_number = format_money_number(total_value)
            total_words = money_to_vietnamese_words(total_value)

        payload = ContractCreate(
            ngay_lap_hop_dong=date.fromisoformat(ngay_lap_hop_dong),
            so_hop_dong_4=so_hop_dong_4,
            linh_vuc=linh_vuc_value,
            don_vi_ten=_clean_opt(don_vi_ten),
            don_vi_dia_chi=_clean_opt(don_vi_dia_chi),
            don_vi_dien_thoai=normalize_multi_phones(don_vi_dien_thoai),
            don_vi_nguoi_dai_dien=_clean_opt(don_vi_nguoi_dai_dien),
            don_vi_chuc_vu=_clean_opt(don_vi_chuc_vu) or "Giám đốc",
            don_vi_mst=_clean_opt(don_vi_mst),
            don_vi_email=normalize_multi_emails(don_vi_email),
            so_CCCD=_clean_opt(so_CCCD),
            ngay_cap_CCCD=_clean_opt(ngay_cap_CCCD),
            nguoi_thuc_hien_email=normalize_multi_emails(nguoi_thuc_hien_email),
            kenh_ten=_clean_opt(kenh_ten),
            kenh_id=channel_id,
            so_tien_chua_GTGT=_clean_opt(so_tien_chua_GTGT) or None,
            thue_percent=_clean_opt(thue_percent) or None,
        )

        year = payload.ngay_lap_hop_dong.year
        contract_no = f"{payload.so_hop_dong_4}/{year}/{REGION_CODE}/{FIELD_CODE}"

        # Legacy money fields kept for compatibility (treat as total)
        money_value = total_value
        money_text = total_text

        # Render DOCX
        out_docx_dir = STORAGE_DOCX_DIR / str(year)
        out_docx_dir.mkdir(parents=True, exist_ok=True)
        filename = build_docx_filename(
            year=year,
            so_hop_dong_4=payload.so_hop_dong_4,
            so_phu_luc=None,
            linh_vuc=linh_vuc_value,
            kenh_ten=payload.kenh_ten or "",
        )
        out_docx_path = out_docx_dir / filename
        if out_docx_path.exists():
            stem = out_docx_path.stem
            out_docx_path = out_docx_dir / f"{stem}_{date.today().strftime('%Y%m%d')}.docx"

        context = {
            "contract_no": contract_no,
            # Alias keys to match user-defined <...> markers in the sample/template
            "so_hop_dong": contract_no,
            "linh_vuc": linh_vuc_value,
            **date_parts(payload.ngay_lap_hop_dong),
            # Distinct marker for contract signing date (for future annex/doc upgrades)
            "ngay_ky_hop_dong": f"{payload.ngay_lap_hop_dong.day:02d}",
            "ngay_ky_hop_dong_day_du": payload.ngay_lap_hop_dong.strftime("%d/%m/%Y"),
            # Template aliases for contract date
            "thang_ky_hop_dong": f"{payload.ngay_lap_hop_dong.month:02d}",
            "nam_ky_hop_dong": f"{payload.ngay_lap_hop_dong.year}",
            "don_vi_ten": payload.don_vi_ten,
            "don_vi_dia_chi": payload.don_vi_dia_chi,
            "don_vi_dien_thoai": payload.don_vi_dien_thoai,
            "don_vi_nguoi_dai_dien": payload.don_vi_nguoi_dai_dien,
            "don_vi_chuc_vu": payload.don_vi_chuc_vu,
            "don_vi_mst": payload.don_vi_mst,
            "don_vi_email": payload.don_vi_email,
            "so_CCCD": payload.so_CCCD or "",
            "ngay_cap_CCCD": payload.ngay_cap_CCCD or "",
            "kenh_ten": payload.kenh_ten,
            "kenh_id": payload.kenh_id,
            "nguoi_thuc_hien_email": payload.nguoi_thuc_hien_email or "",
            # In the DOCX template, currency label 'VNĐ' is usually present already.
            # So we pass number-only strings to avoid 'VNĐ VNĐ'.
            "so_tien_nhuan_but": total_number,

            # Extra aliases from <...> markers (support both lowercase and uppercase)
            "TEN_DON_VI": payload.don_vi_ten,
            "ten_don_vi": payload.don_vi_ten,
            "dia_chi": payload.don_vi_dia_chi,
            "so_dien_thoai": payload.don_vi_dien_thoai,
            "NGUOI_DAI_DIEN": payload.don_vi_nguoi_dai_dien,
            "nguoi_dai_dien": payload.don_vi_nguoi_dai_dien,
            "CHUC_VU": payload.don_vi_chuc_vu,
            "chuc_vu": payload.don_vi_chuc_vu,
            "ma_so_thue": payload.don_vi_mst,
            "email": payload.don_vi_email,
            "ten_kenh": payload.kenh_ten,
            "link_kenh": channel_link,
            "so_tien_chua_GTGT": pre_vat_number,
            "so_tien_GTGT": total_number,
            "thue_GTGT": vat_number,
            "so_tien": total_number,
            "so_tien_bang_chu": total_words,
            "thue_percent": str(int(vat_percent_value)) if vat_percent_value else "10",
        }

        render_contract_docx(
            template_path=DOCX_TEMPLATE_PATH,
            output_path=out_docx_path,
            context=context,
        )

        # Export catalogue Excel from template
        out_excel_dir = STORAGE_EXCEL_DIR / str(year)
        out_excel_dir.mkdir(parents=True, exist_ok=True)
        catalogue_name = out_docx_path.with_suffix(".xlsx").name
        out_catalogue_path = out_excel_dir / catalogue_name

        catalogue_context = dict(context)
        catalogue_context["so_hop_dong_day_du"] = contract_no
        catalogue_context["ngay_ky_hop_dong"] = payload.ngay_lap_hop_dong.strftime("%d/%m/%Y")
        export_catalogue_excel(
            template_path=CATALOGUE_TEMPLATE_PATH,
            output_path=out_catalogue_path,
            context=catalogue_context,
            sheet_name="Final",
        )

        # Append Excel
        excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"
        record = ContractRecord(
            contract_no=contract_no,
            contract_year=year,
            ngay_lap_hop_dong=payload.ngay_lap_hop_dong,
            linh_vuc=linh_vuc_value,
            region_code=REGION_CODE,
            field_code=FIELD_CODE,
            don_vi_ten=payload.don_vi_ten,
            don_vi_dia_chi=payload.don_vi_dia_chi,
            don_vi_dien_thoai=payload.don_vi_dien_thoai,
            don_vi_nguoi_dai_dien=payload.don_vi_nguoi_dai_dien,
            don_vi_chuc_vu=payload.don_vi_chuc_vu,
            don_vi_mst=payload.don_vi_mst,
            don_vi_email=normalize_multi_emails(payload.don_vi_email),
            so_CCCD=payload.so_CCCD or "",
            ngay_cap_CCCD=payload.ngay_cap_CCCD or "",
            kenh_ten=payload.kenh_ten,
            kenh_id=payload.kenh_id,
            nguoi_thuc_hien_email=normalize_multi_emails(payload.nguoi_thuc_hien_email or ""),
            so_tien_nhuan_but_value=money_value,
            so_tien_nhuan_but_text=format_money_number(money_value) if money_value is not None else "",
            so_tien_chua_GTGT_value=pre_vat_value,
            so_tien_chua_GTGT_text=format_money_number(pre_vat_value) if pre_vat_value is not None else "",
            thue_percent=vat_percent_value,
            thue_GTGT_value=vat_value,
            thue_GTGT_text=format_money_number(vat_value) if vat_value is not None else "",
            so_tien_value=total_value,
            so_tien_text=format_money_number(total_value) if total_value is not None else "",
            so_tien_bang_chu=total_words,
            docx_path=str(out_docx_path),
        )
        append_contract_row(excel_path=excel_path, record=record)

        return RedirectResponse(
            url=(
                f"/contracts?year={year}"
                f"&download=/download/{year}/{out_docx_path.name}"
                f"&download2=/download_excel/{year}/{out_catalogue_path.name}"
            ),
            status_code=303,
        )

    except Exception as e:
        traceback.print_exc()
        msg = f"{type(e).__name__}: {e}" if str(e) else f"{type(e).__name__}"
        return RedirectResponse(url=f"/documents/new?doc_type=contract&error={msg}", status_code=303)


@app.get("/api/contracts")
def api_contracts_list(year: int | None = None, q: str | None = None):
    y = year or date.today().year
    excel_path = STORAGE_EXCEL_DIR / f"contracts_{y}.xlsx"
    rows = read_contracts(excel_path=excel_path)

    # Filter: only show contracts (annex_no is empty)
    contracts = [r for r in rows if not r.get("annex_no")]

    # Apply search filter if provided
    if q:
        q_lower = q.lower()
        contracts = [
            c for c in contracts
            if q_lower in (c.get("contract_no") or "").lower()
            or q_lower in (c.get("kenh_ten") or "").lower()
        ]

    # Serialize and return only essential fields
    result = []
    for c in contracts:
        result.append({
            "contract_no": c.get("contract_no"),
            "kenh_ten": c.get("kenh_ten"),
            "don_vi_ten": c.get("don_vi_ten"),
            "kenh_id": c.get("kenh_id"),
        })

    return JSONResponse({"contracts": result})


@app.get("/contracts", response_class=HTMLResponse)
def contracts_list(request: Request, year: int | None = None, download: str | None = None, download2: str | None = None):
    y = year or date.today().year
    excel_path = STORAGE_EXCEL_DIR / f"contracts_{y}.xlsx"

    rows = read_contracts(excel_path=excel_path)

    # Filter: only show contracts (annex_no is empty)
    contracts = [r for r in rows if not r.get("annex_no")]

    # Calculate statistics
    total_contracts = len(contracts)
    total_value = 0
    for r in contracts:
        val = r.get("so_tien_value", 0)
        if val:
            try:
                if isinstance(val, str):
                    val = int(val.replace(",", "").replace(".", ""))
                total_value += int(val)
            except (ValueError, AttributeError):
                pass

    # Count contracts with annexes
    all_contract_nos = {r.get("contract_no") for r in contracts}
    annexes = [r for r in rows if r.get("annex_no")]
    contracts_with_annexes = len({r.get("contract_no") for r in annexes if r.get("contract_no") in all_contract_nos})

    # Add download url and annex count
    for r in contracts:
        path = r.get("docx_path")
        if isinstance(path, str) and path.strip():
            p = Path(path)
            if p.exists():
                filename = p.name
                r["download_url"] = f"/download/{y}/{filename}"
            else:
                r["download_url"] = None
        else:
            r["download_url"] = None

        # Count annexes for this contract
        contract_no = r.get("contract_no")
        r["annex_count"] = len([a for a in annexes if a.get("contract_no") == contract_no])

    stats = {
        "total_contracts": total_contracts,
        "total_value": total_value,
        "contracts_with_annexes": contracts_with_annexes,
    }

    return templates.TemplateResponse(
        "contracts_list.html",
        {
            "request": request,
            "title": "Danh sách hợp đồng",
            "year": y,
            "rows": contracts,
            "stats": stats,
            "download": download,
            "download2": download2,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.get("/annexes", response_class=HTMLResponse)
def annexes_list(request: Request, year: int | None = None, download: str | None = None):
    y = year or date.today().year
    excel_path = STORAGE_EXCEL_DIR / f"contracts_{y}.xlsx"

    rows = read_contracts(excel_path=excel_path)

    # Filter: only show annexes (annex_no is not empty)
    annexes = [r for r in rows if r.get("annex_no")]

    # Get base contracts for reference
    contracts = [r for r in rows if not r.get("annex_no")]
    contracts_map = {r.get("contract_no"): r for r in contracts}

    # Calculate statistics
    total_annexes = len(annexes)
    total_value = 0
    for r in annexes:
        val = r.get("so_tien_value", 0)
        if val:
            try:
                if isinstance(val, str):
                    val = int(val.replace(",", "").replace(".", ""))
                total_value += int(val)
            except (ValueError, AttributeError):
                pass

    # Find contract with most annexes
    contract_annex_counts = {}
    for a in annexes:
        contract_no = a.get("contract_no")
        if contract_no:
            contract_annex_counts[contract_no] = contract_annex_counts.get(contract_no, 0) + 1

    most_annexes_contract = None
    most_annexes_count = 0
    if contract_annex_counts:
        most_annexes_contract = max(contract_annex_counts, key=contract_annex_counts.get)
        most_annexes_count = contract_annex_counts[most_annexes_contract]

    # Add download url and parent contract info
    for r in annexes:
        path = r.get("docx_path")
        if isinstance(path, str) and path.strip():
            p = Path(path)
            if p.exists():
                filename = p.name
                r["download_url"] = f"/download/{y}/{filename}"
            else:
                r["download_url"] = None
        else:
            r["download_url"] = None

        # Add parent contract info
        contract_no = r.get("contract_no")
        if contract_no in contracts_map:
            parent = contracts_map[contract_no]
            r["parent_contract"] = {
                "don_vi_ten": parent.get("don_vi_ten", ""),
                "kenh_ten": parent.get("kenh_ten", ""),
                "ngay_lap_hop_dong": parent.get("ngay_lap_hop_dong", ""),
            }
        else:
            r["parent_contract"] = None

    stats = {
        "total_annexes": total_annexes,
        "total_value": total_value,
        "most_annexes_contract": most_annexes_contract,
        "most_annexes_count": most_annexes_count,
        "unique_contracts": len(contract_annex_counts),
    }

    return templates.TemplateResponse(
        "annexes_list.html",
        {
            "request": request,
            "title": "Danh sách phụ lục",
            "year": y,
            "rows": annexes,
            "stats": stats,
            "download": download,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.get("/download/{year}/{filename}")
def download_docx(year: int, filename: str):
    path = STORAGE_DOCX_DIR / str(year) / filename
    if not path.exists():
        return HTMLResponse("Not found", status_code=404)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=filename,
    )


@app.get("/download_excel/{year}/{filename}")
def download_excel(year: int, filename: str):
    path = STORAGE_EXCEL_DIR / str(year) / filename
    if not path.exists():
        return HTMLResponse("Not found", status_code=404)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


def _serialize_for_json(obj):
    """Convert datetime objects to strings for JSON serialization"""
    if isinstance(obj, dict):
        return {k: _serialize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_serialize_for_json(item) for item in obj]
    elif isinstance(obj, (date, datetime)):
        return obj.isoformat()
    else:
        return obj


@app.get("/contracts/{year}/detail")
def get_contract_detail(year: int, contract_no: str):
    excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"
    rows = read_contracts(excel_path=excel_path)

    contract = None
    for r in rows:
        if r.get("contract_no") == contract_no and not r.get("annex_no"):
            contract = r
            break

    if not contract:
        return JSONResponse({"error": f"Không tìm thấy hợp đồng: {contract_no}"}, status_code=404)

    # Get annexes for this contract
    annexes = [r for r in rows if r.get("contract_no") == contract_no and r.get("annex_no")]

    # Format dates for display
    if contract.get("ngay_lap_hop_dong"):
        val = contract["ngay_lap_hop_dong"]
        if isinstance(val, (date, datetime)):
            contract["ngay_lap_hop_dong_display"] = val.strftime("%d/%m/%Y")

    # Serialize datetime objects for JSON
    contract_serialized = _serialize_for_json(contract)
    annexes_serialized = _serialize_for_json(annexes)

    return JSONResponse({
        "contract": contract_serialized,
        "annexes": annexes_serialized,
    })


@app.get("/contracts/{year}/edit", response_class=HTMLResponse)
def edit_contract_form(request: Request, year: int, contract_no: str):
    excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"
    rows = read_contracts(excel_path=excel_path)

    contract = None
    for r in rows:
        if r.get("contract_no") == contract_no and not r.get("annex_no"):
            contract = r
            break

    if not contract:
        return RedirectResponse(url=f"/contracts?year={year}&error=Không tìm thấy hợp đồng", status_code=303)

    # Format date for form input
    ngay_lap = contract.get("ngay_lap_hop_dong")
    if isinstance(ngay_lap, (date, datetime)):
        if isinstance(ngay_lap, datetime):
            ngay_lap = ngay_lap.date()
        contract["ngay_lap_hop_dong"] = ngay_lap.isoformat()

    # Parse so_hop_dong_4 from contract_no
    so_hop_dong_4 = _parse_so_hop_dong_4(contract_no)
    contract["so_hop_dong_4"] = so_hop_dong_4

    return templates.TemplateResponse(
        "contract_edit.html",
        {
            "request": request,
            "title": f"Chỉnh sửa hợp đồng {contract_no}",
            "contract": contract,
            "year": year,
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/contracts/{year}/update")
def update_contract(
    request: Request,
    year: int,
    contract_no: str = Form(...),
    ngay_lap_hop_dong: str = Form(...),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form(""),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form("10"),
):
    try:
        excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"

        # Calculate money
        pre_vat_value = None
        vat_value = None
        total_value = None
        vat_percent_value = None

        if _clean_opt(so_tien_chua_GTGT):
            pre_vat_value = normalize_money_to_int(_clean_opt(so_tien_chua_GTGT))
            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_value = float(pct_raw.replace(",", "."))
            vat_value = int(round(pre_vat_value * vat_percent_value / 100.0))
            total_value = pre_vat_value + vat_value

        channel_id, channel_link = normalize_youtube_channel_input(kenh_id)

        updated_data = {
            "ngay_lap_hop_dong": date.fromisoformat(ngay_lap_hop_dong),
            "don_vi_ten": _clean_opt(don_vi_ten),
            "don_vi_dia_chi": _clean_opt(don_vi_dia_chi),
            "don_vi_dien_thoai": normalize_multi_phones(don_vi_dien_thoai),
            "don_vi_nguoi_dai_dien": _clean_opt(don_vi_nguoi_dai_dien),
            "don_vi_chuc_vu": _clean_opt(don_vi_chuc_vu),
            "don_vi_mst": _clean_opt(don_vi_mst),
            "don_vi_email": normalize_multi_emails(don_vi_email),
            "kenh_ten": _clean_opt(kenh_ten),
            "kenh_id": channel_id,
            "so_tien_chua_GTGT_value": pre_vat_value,
            "so_tien_chua_GTGT_text": format_money_number(pre_vat_value) if pre_vat_value else "",
            "thue_percent": vat_percent_value,
            "thue_GTGT_value": vat_value,
            "thue_GTGT_text": format_money_number(vat_value) if vat_value else "",
            "so_tien_value": total_value,
            "so_tien_text": format_money_number(total_value) if total_value else "",
            "so_tien_nhuan_but_value": total_value,
            "so_tien_nhuan_but_text": format_money_number(total_value) if total_value else "",
            "so_tien_bang_chu": money_to_vietnamese_words(total_value) if total_value else "",
        }

        success = update_contract_row(
            excel_path=excel_path,
            contract_no=contract_no,
            annex_no=None,
            updated_data=updated_data
        )

        if success:
            return RedirectResponse(url=f"/contracts?year={year}", status_code=303)
        else:
            return RedirectResponse(url=f"/contracts?year={year}&error=Update failed", status_code=303)

    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(url=f"/contracts/{year}/edit?contract_no={quote(contract_no)}&error={str(e)}", status_code=303)


@app.post("/contracts/{year}/delete")
def delete_contract(year: int, contract_no: str):
    try:
        excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"

        # Delete associated DOCX file
        rows = read_contracts(excel_path=excel_path)
        for r in rows:
            if r.get("contract_no") == contract_no and not r.get("annex_no"):
                docx_path = r.get("docx_path")
                if docx_path and isinstance(docx_path, str):
                    p = Path(docx_path)
                    if p.exists():
                        p.unlink()
                break

        success = delete_contract_row(excel_path=excel_path, contract_no=contract_no, annex_no=None)

        if success:
            return JSONResponse({"success": True, "message": "Đã xóa hợp đồng"})
        else:
            return JSONResponse({"success": False, "error": "Không tìm thấy hợp đồng"}, status_code=404)

    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/annexes/{year}/delete")
def delete_annex(year: int, contract_no: str, annex_no: str):
    try:
        excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"

        # Delete associated DOCX and Excel files
        rows = read_contracts(excel_path=excel_path)
        for r in rows:
            if r.get("contract_no") == contract_no and r.get("annex_no") == annex_no:
                # Delete DOCX
                docx_path = r.get("docx_path")
                if docx_path and isinstance(docx_path, str):
                    p = Path(docx_path)
                    if p.exists():
                        p.unlink()

                # Delete catalogue Excel if exists
                catalogue_path = r.get("catalogue_path")
                if catalogue_path and isinstance(catalogue_path, str):
                    p = Path(catalogue_path)
                    if p.exists():
                        p.unlink()
                break

        success = delete_contract_row(excel_path=excel_path, contract_no=contract_no, annex_no=annex_no)

        if success:
            return JSONResponse({"success": True, "message": f"Đã xóa phụ lục {annex_no}"})
        else:
            return JSONResponse({"success": False, "error": "Không tìm thấy phụ lục"}, status_code=404)

    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.get("/annexes/new", response_class=HTMLResponse)
def annex_form(request: Request, year: int | None = None, contract_no: str | None = None, error: str | None = None):
    y = year or date.today().year
    url = f"/documents/new?doc_type=annex&year={y}"
    if contract_no:
        url += f"&contract_no={contract_no}"
    if error:
        url += f"&error={error}"
    return RedirectResponse(url=url)


def _parse_so_hop_dong_4(contract_no: str) -> str:
    # Expected: 0001/2025/HĐQTGAN-PN/MR
    if not contract_no:
        return ""
    parts = contract_no.split("/")
    return parts[0] if parts else ""


@app.post("/annexes")
def create_annex(
    request: Request,
    contract_no: str = Form(...),
    annex_no: str = Form(""),
    ngay_ky_hop_dong: str = Form(""),
    ngay_ky_phu_luc: str = Form(...),
    linh_vuc: str = Form(""),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form(""),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    so_CCCD: str = Form(""),
    ngay_cap_CCCD: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    nguoi_thuc_hien_email: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form("10"),
):
    try:
        so_phu_luc = annex_no.strip() or None

        # Find contract record first (by year derived from contract_no)
        year = None
        parts = contract_no.split("/")
        if len(parts) >= 2 and parts[1].isdigit():
            year = int(parts[1])
        else:
            year = date.today().year

        contracts = read_contracts(excel_path=STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx")
        contract_row: dict | None = None
        for r in contracts:
            if r.get("contract_no") == contract_no:
                contract_row = r
                break

        # Get contract signing date (from form or contract record)
        if ngay_ky_hop_dong and ngay_ky_hop_dong.strip():
            contract_date = date.fromisoformat(ngay_ky_hop_dong)
        elif contract_row and contract_row.get("ngay_lap_hop_dong"):
            contract_date_value = contract_row["ngay_lap_hop_dong"]
            if isinstance(contract_date_value, date):
                if isinstance(contract_date_value, datetime):
                    contract_date = contract_date_value.date()
                else:
                    contract_date = contract_date_value
            elif isinstance(contract_date_value, str):
                contract_date = date.fromisoformat(contract_date_value)
            else:
                return templates.TemplateResponse(
                    "annex_form.html",
                    {
                        "request": request,
                        "title": "Tạo phụ lục",
                        "error": "Không tìm thấy ngày ký hợp đồng",
                        "contracts": contracts,
                        "preview": {},
                        "year": year,
                        "today": date.today().isoformat(),
                        "selected_contract_no": contract_no,
                        "breadcrumbs": get_breadcrumbs(request.url.path),
                    },
                    status_code=400,
                )
        else:
            return templates.TemplateResponse(
                "annex_form.html",
                {
                    "request": request,
                    "title": "Tạo phụ lục",
                    "error": "Vui lòng nhập ngày ký hợp đồng hoặc chọn hợp đồng có sẵn",
                    "contracts": contracts,
                    "preview": {},
                    "year": year,
                    "today": date.today().isoformat(),
                    "selected_contract_no": contract_no,
                    "breadcrumbs": get_breadcrumbs(request.url.path),
                },
                status_code=400,
            )

        contract_date_parts = {
            "ngay_ky_hop_dong": f"{contract_date.day:02d}",
            "thang_ky_hop_dong": f"{contract_date.month:02d}",
            "nam_ky_hop_dong": f"{contract_date.year}",
            "so_hop_dong_day_du": contract_no,
        }

        # Parse annex signing date
        annex_date = date.fromisoformat(ngay_ky_phu_luc)
        annex_date_parts = {
            "ngay_ky_phu_luc": f"{annex_date.day:02d}",
            "thang_ky_phu_luc": f"{annex_date.month:02d}",
            "nam_ky_phu_luc": f"{annex_date.year}",
        }

        # Prefer user overrides; fallback to contract values (if found)
        linh_vuc_value = _clean_opt(linh_vuc) or (contract_row.get("linh_vuc") if contract_row else "") or FIELD_NAME
        don_vi_ten_value = _clean_opt(don_vi_ten) or (contract_row.get("don_vi_ten") if contract_row else "") or ""
        don_vi_dia_chi_value = _clean_opt(don_vi_dia_chi) or (contract_row.get("don_vi_dia_chi") if contract_row else "") or ""
        don_vi_dien_thoai_value = normalize_multi_phones(
            _clean_opt(don_vi_dien_thoai) or (contract_row.get("don_vi_dien_thoai") if contract_row else "") or ""
        )
        don_vi_nguoi_dai_dien_value = _clean_opt(don_vi_nguoi_dai_dien) or (contract_row.get("don_vi_nguoi_dai_dien") if contract_row else "") or ""
        don_vi_chuc_vu_value = _clean_opt(don_vi_chuc_vu) or (contract_row.get("don_vi_chuc_vu") if contract_row else "") or "Giám đốc"
        don_vi_mst_value = _clean_opt(don_vi_mst) or (contract_row.get("don_vi_mst") if contract_row else "") or ""
        don_vi_email_value = normalize_multi_emails(
            _clean_opt(don_vi_email) or (contract_row.get("don_vi_email") if contract_row else "") or ""
        )
        kenh_ten_value = _clean_opt(kenh_ten) or (contract_row.get("kenh_ten") if contract_row else "") or ""

        channel_id_value_raw = _clean_opt(kenh_id) or (contract_row.get("kenh_id") if contract_row else "") or ""
        channel_id_value, channel_link_value = normalize_youtube_channel_input(channel_id_value_raw)

        # Calculate money with VAT
        pre_vat_value = 0
        vat_value = 0
        total_value = 0
        pre_vat_number = ""
        vat_number = ""
        total_number = ""
        total_words = ""

        if _clean_opt(so_tien_chua_GTGT):
            pre_vat_value = normalize_money_to_int(_clean_opt(so_tien_chua_GTGT))
            pre_vat_number = format_money_number(pre_vat_value)

            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_value = float(pct_raw.replace(",", "."))
            if vat_percent_value < 0:
                raise ValueError("Thuế GTGT không hợp lệ")

            vat_value = int(round(pre_vat_value * vat_percent_value / 100.0))
            vat_number = format_money_number(vat_value)

            total_value = pre_vat_value + vat_value
            total_number = format_money_number(total_value)
            total_words = money_to_vietnamese_words(total_value)

        # Build context reusing known placeholders
        context = {
            "contract_no": contract_no,
            "so_hop_dong": contract_no,
            "so_hop_dong_day_du": contract_no,
            "so_phu_luc": so_phu_luc or "",
            "linh_vuc": linh_vuc_value,
            **contract_date_parts,
            **annex_date_parts,
            "don_vi_ten": don_vi_ten_value,
            "don_vi_dia_chi": don_vi_dia_chi_value,
            "don_vi_dien_thoai": don_vi_dien_thoai_value,
            "don_vi_nguoi_dai_dien": don_vi_nguoi_dai_dien_value,
            "don_vi_chuc_vu": don_vi_chuc_vu_value,
            "don_vi_mst": don_vi_mst_value,
            "don_vi_email": don_vi_email_value,
            "so_CCCD": so_CCCD or "",
            "ngay_cap_CCCD": ngay_cap_CCCD or "",
            "kenh_ten": kenh_ten_value,
            "kenh_id": channel_id_value,
            "nguoi_thuc_hien_email": nguoi_thuc_hien_email or "",
            "so_tien_nhuan_but": total_number,
            "so_tien_chua_GTGT": pre_vat_number,
            "thue_GTGT": vat_number,
            "so_tien_GTGT": total_number,
            "so_tien": total_number,
            "so_tien_bang_chu": total_words,
            "thue_percent": str(int(vat_percent_value)) if vat_percent_value else "10",
            "TEN_DON_VI": don_vi_ten_value,
            "ten_don_vi": don_vi_ten_value,
            "dia_chi": don_vi_dia_chi_value,
            "so_dien_thoai": don_vi_dien_thoai_value,
            "NGUOI_DAI_DIEN": don_vi_nguoi_dai_dien_value,
            "nguoi_dai_dien": don_vi_nguoi_dai_dien_value,
            "CHUC_VU": don_vi_chuc_vu_value,
            "chuc_vu": don_vi_chuc_vu_value,
            "ma_so_thue": don_vi_mst_value,
            "email": don_vi_email_value,
            "ten_kenh": kenh_ten_value,
            "link_kenh": channel_link_value,
            "so_tien_nhuan_but": total_number,
            "so_tien_chua_GTGT": pre_vat_number,
            "thue_GTGT": vat_number,
            "so_tien_GTGT": total_number,
            "so_tien": total_number,
            "so_tien_bang_chu": total_words,
            "thue_percent": str(int(vat_percent_value)) if vat_percent_value else "10",
        }

        # Save DOCX
        out_docx_dir = STORAGE_DOCX_DIR / str(year)
        out_docx_dir.mkdir(parents=True, exist_ok=True)
        filename = build_docx_filename(
            year=year,
            so_hop_dong_4=_parse_so_hop_dong_4(contract_no),
            so_phu_luc=so_phu_luc,
            linh_vuc=linh_vuc_value,
            kenh_ten=kenh_ten_value,
        )
        out_docx_path = out_docx_dir / filename
        if out_docx_path.exists():
            stem = out_docx_path.stem
            out_docx_path = out_docx_dir / f"{stem}_{date.today().strftime('%Y%m%d')}.docx"

        render_contract_docx(template_path=ANNEX_TEMPLATE_PATH, output_path=out_docx_path, context=context)

        # Export annex catalogue Excel from template
        out_excel_dir = STORAGE_EXCEL_DIR / str(year)
        out_excel_dir.mkdir(parents=True, exist_ok=True)
        catalogue_name = out_docx_path.with_suffix(".xlsx").name
        out_catalogue_path = out_excel_dir / catalogue_name

        catalogue_context = dict(context)
        catalogue_context["so_hop_dong_day_du"] = contract_no
        catalogue_context["ngay_ky_hop_dong"] = contract_date.strftime("%d/%m/%Y")
        catalogue_context["ngay_ky_phu_luc"] = annex_date.strftime("%d/%m/%Y")
        export_catalogue_excel(
            template_path=ANNEX_CATALOGUE_TEMPLATE_PATH,
            output_path=out_catalogue_path,
            context=catalogue_context,
            sheet_name="Final",
        )

        # Append to contracts Excel with annex_no filled
        contracts_excel_path = STORAGE_EXCEL_DIR / f"contracts_{year}.xlsx"

        # Determine VAT percent value
        vat_percent_final = None
        if _clean_opt(so_tien_chua_GTGT):
            pct_raw = _clean_opt(thue_percent) or "10"
            vat_percent_final = float(pct_raw.replace(",", "."))

        annex_record = ContractRecord(
            contract_no=contract_no,
            contract_year=year,
            annex_no=so_phu_luc,
            ngay_lap_hop_dong=annex_date,
            linh_vuc=linh_vuc_value,
            region_code=REGION_CODE,
            field_code=FIELD_CODE,
            don_vi_ten=don_vi_ten_value,
            don_vi_dia_chi=don_vi_dia_chi_value,
            don_vi_dien_thoai=don_vi_dien_thoai_value,
            don_vi_nguoi_dai_dien=don_vi_nguoi_dai_dien_value,
            don_vi_chuc_vu=don_vi_chuc_vu_value,
            don_vi_mst=don_vi_mst_value,
            don_vi_email=don_vi_email_value,
            kenh_ten=kenh_ten_value,
            kenh_id=channel_id_value,
            so_tien_nhuan_but_value=total_value if total_value else None,
            so_tien_nhuan_but_text=format_money_vnd(total_value) if total_value else None,
            so_tien_chua_GTGT_value=pre_vat_value if pre_vat_value else None,
            so_tien_chua_GTGT_text=format_money_vnd(pre_vat_value) if pre_vat_value else None,
            thue_percent=vat_percent_final,
            thue_GTGT_value=vat_value if vat_value else None,
            thue_GTGT_text=format_money_vnd(vat_value) if vat_value else None,
            so_tien_value=total_value if total_value else None,
            so_tien_text=format_money_vnd(total_value) if total_value else None,
            so_tien_bang_chu=total_words if total_words else None,
            docx_path=str(out_docx_path),
        )
        append_contract_row(excel_path=contracts_excel_path, record=annex_record)

        return RedirectResponse(
            url=(
                f"/contracts?year={year}"
                f"&download=/download/{year}/{out_docx_path.name}"
                f"&download2=/download_excel/{year}/{out_catalogue_path.name}"
            ),
            status_code=303,
        )
    except Exception as e:
        return RedirectResponse(url=f"/documents/new?doc_type=annex&error={str(e)}", status_code=303)


@app.get("/documents/new", response_class=HTMLResponse)
def document_form_unified(
    request: Request,
    doc_type: str | None = None,
    year: int | None = None,
    contract_no: str | None = None,
    error: str | None = None,
):
    y = year or date.today().year
    contracts = read_contracts(excel_path=STORAGE_EXCEL_DIR / f"contracts_{y}.xlsx")

    # Filter: only show base contracts (no annex_no) for dropdown
    contracts = [r for r in contracts if not r.get("annex_no")]

    preview: dict = {}
    if contract_no and doc_type == "annex":
        for r in contracts:
            if r.get("contract_no") == contract_no:
                preview = r.copy()
                if "ngay_lap_hop_dong" in preview:
                    val = preview["ngay_lap_hop_dong"]
                    if isinstance(val, date):
                        if isinstance(val, datetime):
                            val = val.date()
                        preview["ngay_lap_hop_dong"] = val.isoformat()
                break

    return templates.TemplateResponse(
        "document_form.html",
        {
            "request": request,
            "title": "Tạo tài liệu",
            "error": error,
            "doc_type": doc_type or "contract",
            "contracts": contracts,
            "preview": preview,
            "year": y,
            "today": date.today().isoformat(),
            "selected_contract_no": contract_no or "",
            "breadcrumbs": get_breadcrumbs(request.url.path),
        },
    )


@app.post("/documents")
def create_document_unified(
    request: Request,
    doc_type: str = Form(...),
    ngay_lap_hop_dong: str = Form(""),
    so_hop_dong_4: str = Form(""),
    contract_no: str = Form(""),
    annex_no: str = Form(""),
    ngay_ky_hop_dong: str = Form(""),
    ngay_ky_phu_luc: str = Form(""),
    linh_vuc: str = Form(""),
    don_vi_ten: str = Form(""),
    don_vi_dia_chi: str = Form(""),
    don_vi_dien_thoai: str = Form(""),
    don_vi_nguoi_dai_dien: str = Form(""),
    don_vi_chuc_vu: str = Form(""),
    don_vi_mst: str = Form(""),
    don_vi_email: str = Form(""),
    so_CCCD: str = Form(""),
    ngay_cap_CCCD: str = Form(""),
    nguoi_thuc_hien_email: str = Form(""),
    kenh_ten: str = Form(""),
    kenh_id: str = Form(""),
    so_tien_chua_GTGT: str = Form(""),
    thue_percent: str = Form("10"),
):
    if doc_type == "contract":
        # Redirect to contract creation
        return create_contract(
            request=request,
            ngay_lap_hop_dong=ngay_lap_hop_dong,
            so_hop_dong_4=so_hop_dong_4,
            linh_vuc=linh_vuc or FIELD_NAME,
            don_vi_ten=don_vi_ten,
            don_vi_dia_chi=don_vi_dia_chi,
            don_vi_dien_thoai=don_vi_dien_thoai,
            don_vi_nguoi_dai_dien=don_vi_nguoi_dai_dien,
            don_vi_chuc_vu=don_vi_chuc_vu,
            don_vi_mst=don_vi_mst,
            don_vi_email=don_vi_email,
            so_CCCD=so_CCCD,
            ngay_cap_CCCD=ngay_cap_CCCD,
            nguoi_thuc_hien_email=nguoi_thuc_hien_email,
            kenh_ten=kenh_ten,
            kenh_id=kenh_id,
            so_tien_chua_GTGT=so_tien_chua_GTGT,
            thue_percent=thue_percent,
        )
    elif doc_type == "annex":
        # Redirect to annex creation
        return create_annex(
            request=request,
            contract_no=contract_no,
            annex_no=annex_no,
            ngay_ky_hop_dong=ngay_ky_hop_dong,
            ngay_ky_phu_luc=ngay_ky_phu_luc,
            linh_vuc=linh_vuc,
            don_vi_ten=don_vi_ten,
            don_vi_dia_chi=don_vi_dia_chi,
            don_vi_dien_thoai=don_vi_dien_thoai,
            don_vi_nguoi_dai_dien=don_vi_nguoi_dai_dien,
            don_vi_chuc_vu=don_vi_chuc_vu,
            don_vi_mst=don_vi_mst,
            don_vi_email=don_vi_email,
            so_CCCD=so_CCCD,
            ngay_cap_CCCD=ngay_cap_CCCD,
            kenh_ten=kenh_ten,
            kenh_id=kenh_id,
            nguoi_thuc_hien_email=nguoi_thuc_hien_email,
            so_tien_chua_GTGT=so_tien_chua_GTGT,
            thue_percent=thue_percent,
        )
    else:
        return RedirectResponse(url="/documents/new?error=Invalid document type", status_code=303)
