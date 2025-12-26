from __future__ import annotations

from dataclasses import asdict
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.models import ContractRecord
from app.services.backup import BackupManager

backup_manager = BackupManager(storage_dir=Path("storage/excel"))


HEADERS = [
    "contract_no",
    "contract_year",
    "annex_no",
    "ngay_lap_hop_dong",
    "linh_vuc",
    "region_code",
    "field_code",
    "don_vi_ten",
    "don_vi_dia_chi",
    "don_vi_dien_thoai",
    "don_vi_nguoi_dai_dien",
    "don_vi_chuc_vu",
    "don_vi_mst",
    "don_vi_email",

    "so_CCCD",
    "ngay_cap_CCCD",

    "kenh_ten",
    "kenh_id",

    "nguoi_thuc_hien_email",

    "so_tien_nhuan_but_value",
    "so_tien_nhuan_but_text",
    "so_tien_chua_GTGT_value",
    "so_tien_chua_GTGT_text",
    "thue_percent",
    "thue_GTGT_value",
    "thue_GTGT_text",
    "so_tien_value",
    "so_tien_text",
    "so_tien_bang_chu",
    "docx_path",
]


WORKS_HEADERS = [
    "year",
    "contract_no",
    "annex_no",
    "ngay_ky_hop_dong",
    "ngay_ky_phu_luc",
    "nguoi_thuc_hien",
    "ten_kenh",
    "id_channel",
    "link_kenh",
    "stt",
    "id_link",
    "youtube_url",
    "id_work",
    "musical_work",
    "author",
    "composer",
    "lyricist",
    "time_range",
    "duration",
    "effective_date",
    "expiration_date",
    "usage_type",
    "royalty_rate",
    "note",
    "imported_at",
]


_WORKS_FONT = Font(name="Times New Roman", size=12)


def _apply_works_font(ws, *, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).font = _WORKS_FONT


def _rebuild_works_workbook(path: Path) -> None:
    backup = path.with_suffix(path.suffix + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
    shutil.copyfile(path, backup)

    wb_old = load_workbook(str(path))
    ws_old = wb_old["Works"] if "Works" in wb_old.sheetnames else wb_old.active

    old_headers: list[str] = []
    if ws_old.max_row >= 1 and ws_old.max_column >= 1:
        old_headers = [ws_old.cell(row=1, column=c).value for c in range(1, ws_old.max_column + 1)]
    old_headers_norm = [h if isinstance(h, str) else "" for h in old_headers]

    # Preserve extra columns (if any) by keeping them at the end
    extra_headers = [h for h in old_headers_norm if h and h not in WORKS_HEADERS]
    final_headers = WORKS_HEADERS + extra_headers

    # Map header -> column index in old
    col_map: dict[str, int] = {}
    for idx, h in enumerate(old_headers_norm, start=1):
        if h and h not in col_map:
            col_map[h] = idx

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Works"
    ws_new.append(final_headers)
    _apply_works_font(ws_new, row=1, max_col=len(final_headers))

    for r in range(2, ws_old.max_row + 1):
        row_dict: dict[str, object] = {}
        for h, c in col_map.items():
            row_dict[h] = ws_old.cell(row=r, column=c).value
        ws_new.append([row_dict.get(h) for h in final_headers])
        _apply_works_font(ws_new, row=ws_new.max_row, max_col=len(final_headers))

    wb_new.save(str(path))


def _ensure_works_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(str(path))
        if "Works" not in wb.sheetnames:
            ws = wb.create_sheet("Works")
            ws.append(WORKS_HEADERS)
            _apply_works_font(ws, row=1, max_col=len(WORKS_HEADERS))
            wb.save(str(path))
            return

        ws = wb["Works"]
        max_col = ws.max_column if ws.max_column and ws.max_column > 0 else 0
        existing_headers_raw: list[str] = []
        if ws.max_row >= 1 and max_col > 0:
            existing_headers_raw = [ws.cell(row=1, column=i).value for i in range(1, max_col + 1)]
        existing_headers = [h if isinstance(h, str) else "" for h in existing_headers_raw]

        # If headers missing or order differs, rebuild to match our logical order + apply font
        if existing_headers[: len(WORKS_HEADERS)] != WORKS_HEADERS:
            wb.close()
            _rebuild_works_workbook(path)
            return

        # Ensure header row font is consistent
        _apply_works_font(ws, row=1, max_col=max(len(WORKS_HEADERS), ws.max_column))
        wb.save(str(path))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Works"
    ws.append(WORKS_HEADERS)
    _apply_works_font(ws, row=1, max_col=len(WORKS_HEADERS))
    wb.save(str(path))


def append_works_rows(*, excel_path: Path, rows: list[dict]) -> None:
    _ensure_works_workbook(excel_path)

    backup_manager.create_auto_backup(excel_path)

    wb = load_workbook(str(excel_path))
    ws = wb["Works"]

    for r in rows:
        ws.append([r.get(h) for h in WORKS_HEADERS])
        _apply_works_font(ws, row=ws.max_row, max_col=len(WORKS_HEADERS))

    wb.save(str(excel_path))


def _rebuild_contracts_workbook(path: Path) -> None:
    backup = path.with_suffix(path.suffix + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
    shutil.copyfile(path, backup)

    wb_old = load_workbook(str(path))
    ws_old = wb_old["Contracts"] if "Contracts" in wb_old.sheetnames else wb_old.active
    rows = list(ws_old.iter_rows(values_only=True))
    if not rows:
        return

    raw_headers = list(rows[0])
    headers: list[str | None] = []
    for h in raw_headers:
        if isinstance(h, str) and h.strip():
            headers.append(h.strip())
        else:
            headers.append(None)

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Contracts"
    ws_new.append(HEADERS)

    for r in rows[1:]:
        if not any(r):
            continue
        row_dict: dict = {}
        for i in range(min(len(headers), len(r))):
            key = headers[i]
            if not key:
                continue
            v = r[i]
            if isinstance(v, str) and key.endswith("_text"):
                v = v.replace("VNĐ", "").replace("VND", "").strip()
            row_dict[key] = v
        ws_new.append([row_dict.get(h) for h in HEADERS])

    wb_new.save(str(path))


def _ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(str(path))
        ws = wb["Contracts"] if "Contracts" in wb.sheetnames else wb.active
        if ws.title != "Contracts":
            ws.title = "Contracts"

        # Read the header row preserving column positions (including None)
        max_col = ws.max_column if ws.max_column and ws.max_column > 0 else 0
        existing_headers = []
        if ws.max_row >= 1 and max_col > 0:
            existing_headers = [ws.cell(row=1, column=i).value for i in range(1, max_col + 1)]

        if not any(existing_headers):
            ws.append(HEADERS)
            wb.save(str(path))
            return

        existing_set = {h for h in existing_headers if isinstance(h, str) and h}
        missing = [h for h in HEADERS if h not in existing_set]
        if missing:
            # Append after the last used column to avoid shifting existing data
            start_col = (ws.max_column if ws.max_column and ws.max_column > 0 else len(existing_headers)) + 1
            for i, h in enumerate(missing):
                ws.cell(row=1, column=start_col + i, value=h)
            wb.save(str(path))

        # If the workbook has been through schema changes, some rows may be misaligned.
        # Rebuild once when we detect the header row isn't already canonical.
        canonical_prefix = existing_headers[: len(HEADERS)]
        canonical_like = [h if isinstance(h, str) else None for h in canonical_prefix] == HEADERS
        if not canonical_like:
            _rebuild_contracts_workbook(path)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Contracts"
    ws.append(HEADERS)
    wb.save(str(path))


def append_contract_row(*, excel_path: Path, record: ContractRecord) -> None:
    _ensure_workbook(excel_path)

    backup_manager.create_auto_backup(excel_path)

    wb = load_workbook(str(excel_path))
    ws = wb["Contracts"]

    row = []
    data = record.model_dump()
    for h in HEADERS:
        v = data.get(h)
        row.append(v)

    ws.append(row)

    # Set date format for ngay_lap_hop_dong column (dd/mm/yyyy)
    row_num = ws.max_row
    date_col_idx = None
    for idx, header in enumerate(HEADERS, start=1):
        if header == "ngay_lap_hop_dong":
            date_col_idx = idx
            break

    if date_col_idx:
        cell = ws.cell(row=row_num, column=date_col_idx)
        if cell.value:
            cell.number_format = "dd/mm/yyyy"

    wb.save(str(excel_path))


def read_contracts(*, excel_path: Path) -> list[dict]:
    if not excel_path.exists():
        return []

    wb = load_workbook(str(excel_path))
    ws = wb["Contracts"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    raw_headers = list(rows[0])
    headers: list[str | None] = []
    for h in raw_headers:
        if isinstance(h, str) and h.strip():
            headers.append(h.strip())
        else:
            headers.append(None)

    out: list[dict] = []
    for r in rows[1:]:
        if not any(r):
            continue
        row_dict: dict = {}
        for i in range(min(len(headers), len(r))):
            key = headers[i]
            if not key:
                continue
            row_dict[key] = r[i]
        out.append(row_dict)
    return out


def update_contract_row(*, excel_path: Path, contract_no: str, annex_no: str | None, updated_data: dict) -> bool:
    if not excel_path.exists():
        return False

    backup_manager.create_auto_backup(excel_path)

    wb = load_workbook(str(excel_path))
    ws = wb["Contracts"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return False

    raw_headers = list(rows[0])
    headers: list[str | None] = []
    for h in raw_headers:
        if isinstance(h, str) and h.strip():
            headers.append(h.strip())
        else:
            headers.append(None)

    contract_no_idx = None
    annex_no_idx = None
    for i, h in enumerate(headers):
        if h == "contract_no":
            contract_no_idx = i
        if h == "annex_no":
            annex_no_idx = i

    if contract_no_idx is None:
        return False

    found = False
    for row_idx in range(2, ws.max_row + 1):
        row_contract_no = ws.cell(row=row_idx, column=contract_no_idx + 1).value
        row_annex_no = ws.cell(row=row_idx, column=annex_no_idx + 1).value if annex_no_idx is not None else None

        if row_contract_no == contract_no:
            if annex_no is None:
                if not row_annex_no:
                    found = True
            else:
                if row_annex_no == annex_no:
                    found = True

            if found:
                for header_idx, header in enumerate(headers):
                    if header and header in updated_data:
                        ws.cell(row=row_idx, column=header_idx + 1, value=updated_data[header])
                break

    if found:
        wb.save(str(excel_path))
    return found


def delete_contract_row(*, excel_path: Path, contract_no: str, annex_no: str | None = None) -> bool:
    if not excel_path.exists():
        return False

    backup_manager.create_auto_backup(excel_path)

    wb = load_workbook(str(excel_path))
    ws = wb["Contracts"]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return False

    raw_headers = list(rows[0])
    headers: list[str | None] = []
    for h in raw_headers:
        if isinstance(h, str) and h.strip():
            headers.append(h.strip())
        else:
            headers.append(None)

    contract_no_idx = None
    annex_no_idx = None
    for i, h in enumerate(headers):
        if h == "contract_no":
            contract_no_idx = i
        if h == "annex_no":
            annex_no_idx = i

    if contract_no_idx is None:
        return False

    row_to_delete = None
    for row_idx in range(2, ws.max_row + 1):
        row_contract_no = ws.cell(row=row_idx, column=contract_no_idx + 1).value
        row_annex_no = ws.cell(row=row_idx, column=annex_no_idx + 1).value if annex_no_idx is not None else None

        if row_contract_no == contract_no:
            if annex_no is None:
                if not row_annex_no:
                    row_to_delete = row_idx
                    break
            else:
                if row_annex_no == annex_no:
                    row_to_delete = row_idx
                    break

    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(str(excel_path))
        return True
    return False


def export_catalogue_excel(*, template_path: Path, output_path: Path, context: dict, sheet_name: str = "Final") -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(str(output_path))
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    def replace_placeholders(text: str) -> str:
        def repl(m: re.Match[str]) -> str:
            key = m.group(1).strip()
            v = context.get(key, "")
            return "" if v is None else str(v)

        return re.sub(r"<\s*([^<>\s]+)\s*>", repl, text)

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "<" in v and ">" in v:
                cell.value = replace_placeholders(v)

    wb.save(str(output_path))
