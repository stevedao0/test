from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


HEADERS = [
    "contract_no",
    "annex_no",
    "ngay_ky_phu_luc",
    "created_at",
    "docx_path",
]


def _ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(str(path))
        ws = wb["Annexes"] if "Annexes" in wb.sheetnames else wb.active
        if ws.title != "Annexes":
            ws.title = "Annexes"
        if ws.max_row < 1:
            ws.append(HEADERS)
        else:
            existing_headers = [c.value for c in ws[1]]
            if existing_headers[: len(HEADERS)] != HEADERS:
                ws.delete_rows(1)
                ws.insert_rows(1)
                ws.append(HEADERS)
        wb.save(str(path))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Annexes"
    ws.append(HEADERS)
    wb.save(str(path))


def append_annex_row(*, excel_path: Path, record: dict) -> None:
    _ensure_workbook(excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Annexes"]
    ws.append([record.get(h) for h in HEADERS])

    # Set date format for date columns (dd/mm/yyyy)
    row_num = ws.max_row
    for idx, header in enumerate(HEADERS, start=1):
        if header in ("ngay_ky_phu_luc", "created_at"):
            cell = ws.cell(row=row_num, column=idx)
            if cell.value:
                cell.number_format = "dd/mm/yyyy"

    wb.save(str(excel_path))


def read_annexes(*, excel_path: Path) -> list[dict]:
    if not excel_path.exists():
        return []

    wb = load_workbook(str(excel_path))
    ws = wb["Annexes"] if "Annexes" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    raw_headers = list(rows[0])
    headers: list[str | None] = []
    for h in raw_headers:
        if isinstance(h, str) and h.strip():
            headers.append(h.strip())
        else:
            headers.append(None)

    out: list[dict] = []
    for r in rows[1:]:
        if not any(r):
            continue
        row_dict: dict = {}
        for i in range(min(len(headers), len(r))):
            key = headers[i]
            if not key:
                continue
            row_dict[key] = r[i]
        out.append(row_dict)

    return out
